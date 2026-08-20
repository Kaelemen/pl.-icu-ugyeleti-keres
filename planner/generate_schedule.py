# -*- coding: utf-8 -*-
"""
ICU ügyeleti beosztás generátor - szabály-adatbázisból (szabalyok.json) és
hónap-specifikus kívánság-adatbázisból (kivansagok_ÉÉÉÉ_HH.json) dolgozik.
Használat: python3 generate_schedule.py [kivansagok_fajl.json] [sablon.xlsx] [kimenet.xlsx]
"""
import sys
import json
import math
import datetime
import random
import openpyxl
from openpyxl.styles import PatternFill, Font as XLFont

import os
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SZABALYOK_PATH = os.path.join(SCRIPT_DIR, "szabalyok.json")
KIVANSAGOK_PATH = sys.argv[1] if len(sys.argv) > 1 else os.path.join(SCRIPT_DIR, "kivansagok_2026_08.json")
SABLON_PATH = sys.argv[2] if len(sys.argv) > 2 else "ICU_ugyeleti_beosztas.xlsx"
KIMENET_PATH = sys.argv[3] if len(sys.argv) > 3 else "ICU_ugyeleti_beosztas_probabeosztas.xlsx"
SEED = int(sys.argv[4]) if len(sys.argv) > 4 else 0
rng = random.Random(SEED)

with open(SZABALYOK_PATH, encoding="utf-8") as f:
    SZAB = json.load(f)
with open(KIVANSAGOK_PATH, encoding="utf-8") as f:
    KIV = json.load(f)
MINDENKEPPEN_SZERETNE = KIV.get("mindenkeppen_szeretne", {})  # nev -> [napok], amiket mindenképp szeretne ügyeletben
ELOZO_HONAP_LELEPOK = set(KIV.get("elozo_honap_lelepok", []))  # kik voltak ügyeletben az előző hónap utolsó napján

# Ha a kívánság-fájl tartalmaz "dolgozok" listát (a webes admin felület Dolgozók
# kezelése szekciójából jön), az felülírja a szabalyok.json-ban lévő törzsadatot -
# így az admin felületen felvett/törölt/módosított dolgozók azonnal érvényesülnek,
# nem kell külön kézzel frissíteni a szabalyok.json-t is.
if KIV.get("dolgozok"):
    SZAB["dolgozok"] = KIV["dolgozok"]
    print(f"Dolgozói törzsadat felülírva a kívánság-fájlból ({len(KIV['dolgozok'])} fő).")

YEAR, MONTH = KIV["ev"], KIV["honap"]
first_day = datetime.date(YEAR, MONTH, 1)
next_month = datetime.date(YEAR, MONTH + 1, 1) if MONTH != 12 else datetime.date(YEAR + 1, 1, 1)
num_days = (next_month - first_day).days

WEEKDAY_HU = {"hetfo": 0, "kedd": 1, "szerda": 2, "csutortok": 3, "pentek": 4, "szombat": 5, "vasarnap": 6}

# ---------------------------------------------------------------------------
# Dolgozók betöltése a szabalyok.json-ból
# ---------------------------------------------------------------------------
staff = []
for d in SZAB["dolgozok"]:
    staff.append((d["nev"], d["kategoria"], d["napi_munkaido"], KIV.get("kert_ugyeletszam", {}).get(d["nev"]), d["tipus"]))
staff_order_all = [d["nev"] for d in SZAB["dolgozok"]]
keret_tipus_map = {d["nev"]: ("Havi" if d["szerzodes_tipus"] == "Részmunkaidő - havi órakeret" else "Napi")
                    for d in SZAB["dolgozok"]}
havi_oraszam_map = {d["nev"]: d["havi_oraszam"] for d in SZAB["dolgozok"] if d["havi_oraszam"]}
HAVI_KERETESEK = {nev for nev, kt in keret_tipus_map.items() if kt == "Havi"}
REZIDENSEK = {d["nev"] for d in SZAB["dolgozok"] if d["tipus"] == "Rezidens"}
KULSOS_GYAKORLATON = set(KIV.get("kulsos_gyakorlaton", []))  # rezidensek, akik ebben a hónapban külsős gyakorlaton vannak

# 3 havi kiegyenlítés: az előző hónap túlórája (Google Drive-ból automatikusan beolvasva,
# vagy admin által kézzel megadva) csökkenti (ha pozitív), illetve növeli (ha negatív, azaz
# hiány volt) a jelen hónap kötelező óraszám-keretét - így törekszünk rá, hogy 3 hónap alatt
# nullázódjon a túlóra/hiány.
ELOZO_HONAP_TULORA = KIV.get("elozo_honap_tulora", {})
for _nev, _tulora in ELOZO_HONAP_TULORA.items():
    if _nev in havi_oraszam_map:
        havi_oraszam_map[_nev] = havi_oraszam_map[_nev] - _tulora

def keret_of(name):
    return keret_tipus_map.get(name, "Napi")

ALT = SZAB["altalanos_szabalyok"]
MIN_PIHENO = ALT["min_pihenonap_ket_ugyelet_kozott"]
SZOMBAT_NINCS_STROKE = ALT["szombaton_nincs_stroke"]
MIN_SZAKORVOS = ALT["min_szakorvos_naponta_ugyeletben"]
O1_ALAP = ALT["o1_alapertelmezett_szemely"]
O2_ALAP = ALT["o2_alapertelmezett_szemely"]
MUTO_MIN = ALT["muto_minimum_letszam"]
MUTO_PADLO = ALT["muto_padlo_letszam"]
O1_O2_TILTOTT = set(ALT.get("o1_o2_tiltott_szemelyek", []))

NAPI_KOTELEZO_ORA = SZAB["orakeret_konstansok"]["napi_kotelezo_ora"]
munkanapok_a_honapban = sum(1 for d in range(num_days)
                             if (first_day + datetime.timedelta(days=d)).weekday() < 5)
RESZ_NAPI_ORASZAMOS = {d["nev"]: d["napi_munkaido"] for d in SZAB["dolgozok"]
                        if d["szerzodes_tipus"] == "Részmunkaidő - napi óraszám"}
RESZ_NAPI_KAPACITAS = {nev: napi * munkanapok_a_honapban for nev, napi in RESZ_NAPI_ORASZAMOS.items()}
for _nev, _tulora in ELOZO_HONAP_TULORA.items():
    if _nev in RESZ_NAPI_KAPACITAS:
        RESZ_NAPI_KAPACITAS[_nev] = RESZ_NAPI_KAPACITAS[_nev] - _tulora

# "Fél állás" (napi 4 órás) dolgozóknál havonta max. 3 ügyelet lehet, függetlenül attól,
# hogy az óra-kapacitájuk elméletileg többet is megengedne.
FEL_ALLAS_MAX_UGYELET = 3
FEL_ALLAS_NEVEK = {nev for nev, napi in RESZ_NAPI_ORASZAMOS.items() if napi == 4.0}
kotelezo_ora_used = {nev: 0 for nev in RESZ_NAPI_ORASZAMOS}

# ---------------------------------------------------------------------------
# Kívánságok betöltése (hónap-specifikus), majd a személyi (tartós) szabályok
# alkalmazása dátumokra bontva erre a hónapra
# ---------------------------------------------------------------------------
kivansagok = {name: {"szabadsag": [], "nem": [], "szeret": []} for name in staff_order_all}
for name, v in KIV["kivansagok"].items():
    kivansagok[name] = {"szabadsag": list(v.get("szabadsag", [])),
                         "nem": list(v.get("nem", [])),
                         "szeret": list(v.get("szeret", []))}

# Pillanatkép a TÉNYLEGESEN, kifejezetten beküldött kérésekről - mielőtt bármilyen
# személyi szabály vagy "csak jelölt napokon dolgozik" logika automatikusan kibővítené
# a "nem" listát. Ez a nyomtatható lap színezéséhez kell, hogy csak azt mutassuk pirosnak,
# amit a dolgozó ténylegesen "nem szeretne"-ként jelölt - ne az algoritmus belső,
# levezetett kiegészítéseit.
EREDETI_KIVANSAGOK = {name: {k: list(v) for k, v in p.items()} for name, p in kivansagok.items()}

def days_by_weekday(weekday_idx):
    out = []
    for d in range(num_days):
        if (first_day + datetime.timedelta(days=d)).weekday() == weekday_idx:
            out.append(d + 1)
    return out

CSAK_JELOLT_NAPOKON = set()
RENDES_NAP_CSAK_HETENTE = {}   # name -> set(weekday_idx), csak ezeken a napokon lehet "m"
HETI_FIX_ESEMENY = {}   # name -> {"weekday": int, "kod": str}
RESZMUNKAIDO_TOL = {}   # name -> első nap (int), amitől "rendesen" jelen van
# Azok a napok, amik KIZÁRÓLAG a "tiltott_napok_hetente" (heti ismétlődő ügyelet-tiltás,
# pl. Zöldréti: nem vihet keddi ügyeletet) szabály miatt kerültek a "nem" listába - ez csak
# az ÜGYELETRE vonatkozik, nem szabad a jelenlétét (rendes "m" napját) is blokkolnia, ha
# valaki emellett "csak jelölt napokon dolgozik" típusú (pl. rész-munkaidős) besorolást kap.
CSAK_UGYELET_TILTAS_NAPOK = {}   # name -> set(napok)
PARBAN_TILTOTT = []   # [(nev1, nev2, {kivetel_weekday_idx, ...})]

for szab in SZAB["szemelyi_megkotesek"]:
    if szab["tipus"] == "par_nem_egyutt":
        nevek = szab["nevek"]
        if not all(n in staff_order_all for n in nevek):
            continue  # valamelyik dolgozó törölve lett - a régi szabály kihagyva
        kivetel = {WEEKDAY_HU[w] for w in szab.get("kivetel_hetnapok", [])}
        PARBAN_TILTOTT.append((nevek[0], nevek[1], kivetel))
        continue
    name = szab["nev"]
    if name not in staff_order_all:
        continue  # a dolgozó törölve lett az admin felületen - a régi személyi szabálya kihagyva
    tipus = szab["tipus"]
    if tipus == "tiltott_napok_hetente":
        days = []
        for wd in szab["napok"]:
            days += days_by_weekday(WEEKDAY_HU[wd])
        for d in days:
            if d not in kivansagok[name]["nem"] and d not in kivansagok[name]["szeret"]:
                kivansagok[name]["nem"].append(d)
                CSAK_UGYELET_TILTAS_NAPOK.setdefault(name, set()).add(d)
    elif tipus == "tiltott_kezdes_hetente":
        days = []
        for wd in szab["napok"]:
            days += days_by_weekday(WEEKDAY_HU[wd])
        for d in days:
            if d not in kivansagok[name]["nem"] and d not in kivansagok[name]["szeret"]:
                kivansagok[name]["nem"].append(d)
    elif tipus == "nem_dolgozik_hetente":
        days = []
        for wd in szab["napok"]:
            days += days_by_weekday(WEEKDAY_HU[wd])
        for d in days:
            if d not in kivansagok[name]["nem"] and d not in kivansagok[name]["szeret"]:
                kivansagok[name]["nem"].append(d)
        HETI_FIX_ESEMENY.setdefault(name, {})["nem_dolgozik_weekday"] = [WEEKDAY_HU[w] for w in szab["napok"]]
    elif tipus == "csak_jelolt_napokon_dolgozik":
        CSAK_JELOLT_NAPOKON.add(name)
    elif tipus == "heti_fix_esemeny":
        HETI_FIX_ESEMENY.setdefault(name, {})["kod"] = szab["kod"]
        HETI_FIX_ESEMENY[name]["weekday"] = WEEKDAY_HU[szab["nap"]]
    elif tipus == "rendes_nap_csak_hetente":
        RENDES_NAP_CSAK_HETENTE[name] = {WEEKDAY_HU[w] for w in szab["napok"]}

# Minden "Részmunkaidő - napi óraszám" dolgozó alapból csak a saját jelölt ("Szeretne")
# napjain lehet jelen (bármilyen szerepben) - kivéve, akinek konkrét heti mintája van
# (pl. Korompai Máté: rendes_nap_csak_hetente).
CSAK_JELOLT_NAPOKON_KAPACITAS_MIATT = set()  # ide csak a részmunkaidő-napi-óraszám miatt automatikusan bekerültek
for _d in SZAB["dolgozok"]:
    if _d["szerzodes_tipus"] == "Részmunkaidő - napi óraszám" and _d["nev"] not in RENDES_NAP_CSAK_HETENTE:
        CSAK_JELOLT_NAPOKON.add(_d["nev"])
        CSAK_JELOLT_NAPOKON_KAPACITAS_MIATT.add(_d["nev"])

# Részmunkaidő-periódusok beolvasása ELŐBB kell, mint a "csak jelölt napokon" szigorítás,
# hogy tudjuk: kinek van strukturális időszak-korlátja (pl. Gulya: csak a hónap 2. felében
# dolgozik) - az ilyen embereknél a "jó napok" lista NEM egy szűkítő plusz-korlát, hanem
# csak ügyelet-preferencia; a jelenlétüket kizárólag az időszak-korlát szabja meg, a teljes
# időszakon belül egyébként elérhetők.
for name, v in KIV.get("reszmunkaido_periodusok", {}).items():
    tipus = v.get("tipus")
    erinti = set(v.get("erinti", ["jelenlet"]))
    if tipus == "csak_ettol_a_naptol":
        RESZMUNKAIDO_TOL[name] = {"kezdo_nap": v["kezdo_nap"], "erinti": erinti}
    elif tipus == "csak_eddig_a_napig":
        RESZMUNKAIDO_TOL[name] = {"utolso_nap": v["utolso_nap"], "erinti": erinti}

# "csak jelölt napokon dolgozik": minden más nap (ami nincs szabin/szeretve/nemben) -> Nem szeretne.
# Kivéve: (a) ha valakinek EGYÁLTALÁN NINCS "jó napja" megadva, ÉS ez a megkötése kizárólag a
# részmunkaidő-napi-óraszám kapacitása miatt jött létre automatikusan - akkor nem korlátozzuk,
# hogy a havi kötelező órája teljesíthető legyen. Akinek ez SZEMÉLYI (állandó) szabálya - mint
# Berkes Tíbornak -, annál a hiányzó "jó nap" azt jelenti, hogy ebben a hónapban nincs bent,
# NEM azt, hogy szabadon beosztható; (b) ha valakinek strukturális időszak-korlátja van a
# jelenlétére nézve (pl. Gulya) - annál a "jó napok" lista csak ügyelet-preferencia, nem
# szűkíti tovább a jelenlétét az időszakon belül.
EREDETI_KIFEJEZETT_NEM = {name: set(kivansagok[name]["nem"]) - CSAK_UGYELET_TILTAS_NAPOK.get(name, set())
                           for name in CSAK_JELOLT_NAPOKON}
for name in CSAK_JELOLT_NAPOKON:
    p = kivansagok[name]
    if not p["szeret"] and name in CSAK_JELOLT_NAPOKON_KAPACITAS_MIATT:
        continue
    info = RESZMUNKAIDO_TOL.get(name)
    if info and "jelenlet" in info.get("erinti", ()):
        continue
    covered = set(p["szabadsag"]) | set(p["szeret"]) | set(p["nem"])
    for d in range(1, num_days + 1):
        if d not in covered:
            p["nem"].append(d)

def nap_engedelyezett(name, day_nap_szam, hatas):
    """hatas: 'jelenlet' vagy 'ugyelet' - engedélyezett-e ez a nap erre a hatásra nézve"""
    info = RESZMUNKAIDO_TOL.get(name)
    if not info or hatas not in info.get("erinti", ()):
        return True
    if "kezdo_nap" in info:
        return day_nap_szam >= info["kezdo_nap"]
    if "utolso_nap" in info:
        return day_nap_szam <= info["utolso_nap"]
    return True

prefs = {}
for name, p in kivansagok.items():
    for d in p["szabadsag"]:
        prefs[(name, first_day + datetime.timedelta(days=d - 1))] = "Szabadság"
    for d in p["nem"]:
        prefs[(name, first_day + datetime.timedelta(days=d - 1))] = "Nem szeretne"
    for d in p["szeret"]:
        prefs[(name, first_day + datetime.timedelta(days=d - 1))] = "Szeretne"

def is_szabadsag(name, day_date):
    return prefs.get((name, day_date)) == "Szabadság"

def kotelezo_delta_ha_ma_ugyel(name, day_date):
    delta = NAPI_KOTELEZO_ORA
    if (day_date + datetime.timedelta(days=1)).weekday() < 5:
        delta += RESZ_NAPI_ORASZAMOS.get(name, NAPI_KOTELEZO_ORA)
    return delta

TULLEPES_TURESHATAR = 7  # óra - a havi kötelezőt inkább kicsit túl, mint alul, de max ennyivel lépheti túl

def would_exceed_resz_kapacitas(name, day_date):
    if name not in RESZ_NAPI_ORASZAMOS:
        return False
    delta = kotelezo_delta_ha_ma_ugyel(name, day_date)
    return kotelezo_ora_used[name] + delta > RESZ_NAPI_KAPACITAS[name] + TULLEPES_TURESHATAR

def heti_fix_esemeny_ma(name, day_date):
    info = HETI_FIX_ESEMENY.get(name)
    if info and "kod" in info and day_date.weekday() == info["weekday"]:
        return info["kod"]
    return None

def nem_dolgozik_hetente_ma(name, day_date):
    info = HETI_FIX_ESEMENY.get(name)
    return bool(info and day_date.weekday() in info.get("nem_dolgozik_weekday", []))

def jelenlet_tiltott(name, day_date):
    day_nap = (day_date - first_day).days + 1
    return not nap_engedelyezett(name, day_nap, "jelenlet")

def ugyelet_tiltott(name, day_date):
    day_nap = (day_date - first_day).days + 1
    if jelenlet_tiltott(name, day_date):
        return True  # ha nincs jelen (pl. "csak ettől a naptól" korlát), ügyeletre sem osztható be
    return not nap_engedelyezett(name, day_nap, "ugyelet")

def parban_tiltott_utkozik(name, day_date, today_assigned):
    """Két konkrét személy (pl. Kelemen és Katona) nem lehet egyszerre ügyeletben,
    kivéve a szabályban megjelölt kivétel-napokon (pl. péntek/szombat)."""
    for n1, n2, kivetel in PARBAN_TILTOTT:
        if day_date.weekday() in kivetel:
            continue
        if name == n1 and n2 in today_assigned:
            return True
        if name == n2 and n1 in today_assigned:
            return True
    return False

def fel_allas_tullepne(name, req=None, extra=1):
    """Fél állású (napi 4 órás) dolgozónál havonta alapból max. FEL_ALLAS_MAX_UGYELET
    ügyelet lehet - ez a korlát csak akkor hágható át, ha ő maga kifejezetten többet
    kért (kért ügyeletszám) - ilyenkor a saját kérése lesz az érvényes felső korlát."""
    if name not in FEL_ALLAS_NEVEK:
        return False
    hatarertek = max(FEL_ALLAS_MAX_UGYELET, req) if req else FEL_ALLAS_MAX_UGYELET
    return (assigned_count[name] + extra) > hatarertek

def eligible(cat, duty):
    if cat == "T":
        return False  # T kategória: sosem osztható be semmilyen ügyeletre
    if duty == "Intenzív":
        return cat == "I"
    if duty == "Aneszt":
        return cat in ("I", "A")
    if duty == "Stroke":
        return cat in ("I", "A", "St")
    return False

tipus_of = {name: tipus for name, _, _, _, tipus in staff}
cat_of = {name: cat for name, cat, *_ in staff}
T_KATEGORIA_NEVEK = {name for name, cat in cat_of.items() if cat == "T"}
assigned_count = {name: 0 for name, *_ in staff}
target_weight = {}
for name, _cat, hrs, req, _tipus in staff:
    if req:
        target_weight[name] = req
    elif name in HAVI_KERETESEK and havi_oraszam_map.get(name):
        # Havi keretes dolgozóknál (pl. Pintér Enikő) a célérték a havi órakeretükből
        # számolt szükséges ügyeletszám (24 óra/ügyelet), NEM a napi óradíjuk - különben
        # az arányossági logika túl korán "eleget kapottnak" tekinti őket, és a havi
        # kötelező órájuk nem teljesül.
        target_weight[name] = havi_oraszam_map[name] / 24
    else:
        target_weight[name] = hrs
# Minden ügyeleti napot nyilvántartunk (nem csak az utolsót!) - az elsőbbségi kör miatt
# előfordulhat, hogy egy KÉSŐBBI napra már be van osztva valaki, mielőtt a fő ciklus elér egy
# KORÁBBI naphoz - egyetlen "utolsó dátum" mező ilyenkor felülíródna és elveszne a védelem.
duty_dates = {name: set() for name, *_ in staff}

# Az előző hónap utolsó napján ügyeletben lévők (ELOZO_HONAP_LELEPOK) "virtuális" ügyelet-
# dátumot kapnak (az előző hónap utolsó napja) a duty_dates nyilvántartásban is - enélkül a
# fő ügyelet-kiosztó kör pihenőidő-ellenőrzése nem tudna róla, és előfordulhatna, hogy valaki
# már a hónap 1-2. napján új ügyeletet kapna annak ellenére, hogy ténylegesen lelépő állapotban van.
for _nev in ELOZO_HONAP_LELEPOK:
    if _nev in duty_dates:
        duty_dates[_nev].add(first_day - datetime.timedelta(days=1))

def piheno_utkozik(name, day_date):
    return any(abs((day_date - dd).days) <= MIN_PIHENO for dd in duty_dates[name])


def would_exceed_havi_kvota(name, extra_duties=1):
    if name not in HAVI_KERETESEK:
        return False
    kvota = havi_oraszam_map.get(name)
    if not kvota:
        return False
    return (assigned_count[name] + extra_duties) * 24 > kvota

JAVASOLT_KIVETELEK = []  # [{"nap":, "tipus":, "nev":, "szabaly":}] - amiket csak szabály-áthágással
                          # lehetne kitölteni, admin jóváhagyásra várva
ENGEDELYEZETT_KIVETELEK = KIV.get("engedelyezett_kivetelek", [])  # admin által előzőleg jóváhagyott áthágások

SZABALY_LEIRASOK = {
    "piheno": "pihenőidő (min. 2 nap két ügyelet között)",
    "kapacitas": "rész-munkaidős havi óra-kapacitás túllépése",
    "havi_kvota": "havi keretes óraszám túllépése",
    "nem_szeretne": "kifejezetten jelezte, hogy nem szeretne dolgozni aznap",
}


def kivetel_jeloltet_keres(duty, day_date, today_assigned):
    """Megkeresi, ki tudná betölteni a szerepet, ha PONTOSAN EGY konkrét szabályt
    megsértenénk - a kategória-egyezés és a szabadság sosem hágható át, ezekkel
    sosem próbálkozunk. A leggyengébb (legkevésbé súlyos) szabálysértést részesíti
    előnyben."""
    jeloltek_szabalyonkent = {"nem_szeretne": [], "piheno": [], "kapacitas": [], "havi_kvota": []}
    for name, cat, hrs, req, tipus in staff:
        if not eligible(cat, duty):
            continue
        if name in today_assigned:
            continue
        pref = prefs.get((name, day_date))
        if pref == "Szabadság":
            continue  # ez sosem hágható át
        if ugyelet_tiltott(name, day_date):
            continue  # személyi ügyelet-tiltás (pl. heti fix nap) sosem hágható át
        if parban_tiltott_utkozik(name, day_date, today_assigned):
            continue  # páros-tiltás (pl. Kelemen+Katona) sosem hágható át
        if fel_allas_tullepne(name, req):
            continue  # fél állásúaknál a havi 3 ügyelet kemény felső korlát, sosem hágható át
        serult_szabaly = None
        if pref == "Nem szeretne":
            serult_szabaly = "nem_szeretne"
        elif piheno_utkozik(name, day_date):
            serult_szabaly = "piheno"
        elif would_exceed_resz_kapacitas(name, day_date):
            serult_szabaly = "kapacitas"
        elif would_exceed_havi_kvota(name):
            serult_szabaly = "havi_kvota"
        if serult_szabaly is None:
            continue
        ratio = assigned_count[name] / target_weight[name]
        jeloltek_szabalyonkent[serult_szabaly].append((ratio, name))
    for szabaly in ("nem_szeretne", "piheno", "kapacitas", "havi_kvota"):
        if jeloltek_szabalyonkent[szabaly]:
            jeloltek_szabalyonkent[szabaly].sort()
            return szabaly, jeloltek_szabalyonkent[szabaly][0][1]
    return None, None

duty_types = ["Intenzív", "Aneszt", "Stroke"]
schedule = {d: {} for d in range(num_days)}
today_assigned_by_day = {d: set() for d in range(num_days)}

# ---------------------------------------------------------------------------
# ELSŐBBSÉGI KÖR: "mindenképpen szeretném" napok beírása a fő kiosztás előtt.
# Csak akkor íródik be, ha nem ütközik semmilyen kemény szabállyal (pihenőidő,
# havi/rész-munkaidős keret, személyi ügyelet-tiltás, kategória-egyezés, az adott
# napon már ne legyen ügyeletben). A "legalább egy szakorvos" utólagos ellenőrzés
# ugyanúgy lefut ezekre a napokra is, mint bármelyik másikra.
for name, napok in MINDENKEPPEN_SZERETNE.items():
    if name not in tipus_of:
        continue  # ismeretlen név (pl. törölt dolgozó) - kihagyjuk
    cat = next((c for n, c, *_ in staff if n == name), None)
    req_erteke = next((r for n, _, _, r, _ in staff if n == name), None)
    if cat is None:
        continue
    for day_nap in sorted(napok):
        d = day_nap - 1
        if d < 0 or d >= num_days:
            continue
        day_date = first_day + datetime.timedelta(days=d)
        is_saturday_p = day_date.weekday() == 5
        if name in today_assigned_by_day[d]:
            continue  # már be van osztva aznap valamelyik (korábbi) elsőbbségi napja miatt
        jovahagyott = any(k["nap"] == day_nap and k["tipus"] == "mindenkeppen" and k["nev"] == name
                           for k in ENGEDELYEZETT_KIVETELEK)
        if jovahagyott:
            for duty in duty_types:
                if duty == "Stroke" and is_saturday_p and SZOMBAT_NINCS_STROKE:
                    continue
                if schedule[d].get(duty) is not None:
                    continue
                if not eligible(cat, duty):
                    continue
                schedule[d][duty] = name
                today_assigned_by_day[d].add(name)
                assigned_count[name] += 1
                duty_dates[name].add(day_date)
                if name in RESZ_NAPI_ORASZAMOS:
                    kotelezo_ora_used[name] += kotelezo_delta_ha_ma_ugyel(name, day_date)
                break
            continue
        pref = prefs.get((name, day_date))
        if pref == "Szabadság":
            continue  # ez sosem hágható át
        if ugyelet_tiltott(name, day_date):
            continue  # személyi ügyelet-tiltás sosem hágható át
        if parban_tiltott_utkozik(name, day_date, today_assigned_by_day[d]):
            continue  # páros-tiltás (pl. Kelemen+Katona) sosem hágható át
        if fel_allas_tullepne(name, req_erteke):
            continue  # fél állásúaknál alapból max. 3 ügyelet, kivéve ha ő maga többet kért
        serult_szabaly = None
        if pref == "Nem szeretne":
            serult_szabaly = "nem_szeretne"
        elif piheno_utkozik(name, day_date):
            serult_szabaly = "piheno"
        elif would_exceed_havi_kvota(name):
            serult_szabaly = "havi_kvota"
        elif would_exceed_resz_kapacitas(name, day_date):
            serult_szabaly = "kapacitas"
        van_szabad_szerep = any(
            schedule[d].get(duty) is None and eligible(cat, duty)
            and not (duty == "Stroke" and is_saturday_p and SZOMBAT_NINCS_STROKE)
            for duty in duty_types
        )
        if serult_szabaly is not None:
            if van_szabad_szerep:
                JAVASOLT_KIVETELEK.append({
                    "nap": day_nap, "tipus": "mindenkeppen", "nev": name, "szabaly": serult_szabaly,
                    "leiras": f"{name} mindenképp szeretné a(z) {day_nap}. napot, de ehhez át kellene "
                              f"hágni: {SZABALY_LEIRASOK[serult_szabaly]}.",
                })
            continue
        for duty in duty_types:
            if duty == "Stroke" and is_saturday_p and SZOMBAT_NINCS_STROKE:
                continue
            if schedule[d].get(duty) is not None:
                continue  # ezt a szerepet aznap már betöltötte valaki (max. napi keret)
            if not eligible(cat, duty):
                continue
            schedule[d][duty] = name
            today_assigned_by_day[d].add(name)
            assigned_count[name] += 1
            duty_dates[name].add(day_date)
            if name in RESZ_NAPI_ORASZAMOS:
                kotelezo_ora_used[name] += kotelezo_delta_ha_ma_ugyel(name, day_date)
            break

for d in range(num_days):
    day_date = first_day + datetime.timedelta(days=d)
    is_saturday = day_date.weekday() == 5
    today_assigned = today_assigned_by_day[d]
    for duty in duty_types:
        if duty == "Stroke" and is_saturday and SZOMBAT_NINCS_STROKE:
            continue
        if schedule[d].get(duty) is not None:
            continue  # az elsőbbségi kör már betöltötte ezt a szerepet aznap
        candidates = []
        for name, cat, hrs, req, tipus in staff:
            if not eligible(cat, duty):
                continue
            if name in today_assigned:
                continue
            pref = prefs.get((name, day_date))
            if pref in ("Szabadság", "Nem szeretne"):
                continue
            if piheno_utkozik(name, day_date):
                continue
            if would_exceed_havi_kvota(name):
                continue
            if ugyelet_tiltott(name, day_date):
                continue
            if parban_tiltott_utkozik(name, day_date, today_assigned):
                continue
            if fel_allas_tullepne(name, req):
                continue
            if would_exceed_resz_kapacitas(name, day_date):
                continue
            ratio = assigned_count[name] / target_weight[name]
            # Ha valaki kifejezetten kért egy konkrét ügyeletszámot, és még nem érte el -
            # erős elsőbbséget kap, amíg meg nem kapja (a fenti kemény szabályok, kapacitás,
            # pihenőidő stb. továbbra is érvényesek, csak a "kit válasszunk" versenyben nyer).
            kert_meg_nincs_meg = req is not None and req > 0 and assigned_count[name] < req
            bonus = -1.0 if kert_meg_nincs_meg else (-0.3 if pref == "Szeretne" else 0.0)
            jitter = (rng.random() - 0.5) * 0.06
            candidates.append((ratio + bonus + jitter, assigned_count[name], name))
        if not candidates:
            day_nap = d + 1
            jovahagyott_nev = next((k["nev"] for k in ENGEDELYEZETT_KIVETELEK
                                     if k["nap"] == day_nap and k["tipus"] == duty), None)
            if jovahagyott_nev:
                chosen = jovahagyott_nev
                schedule[d][duty] = chosen
                today_assigned.add(chosen)
                assigned_count[chosen] += 1
                duty_dates[chosen].add(day_date)
                if chosen in RESZ_NAPI_ORASZAMOS:
                    kotelezo_ora_used[chosen] += kotelezo_delta_ha_ma_ugyel(chosen, day_date)
                continue
            szabaly, jelolt_nev = kivetel_jeloltet_keres(duty, day_date, today_assigned)
            if jelolt_nev:
                JAVASOLT_KIVETELEK.append({
                    "nap": day_nap, "tipus": duty, "nev": jelolt_nev, "szabaly": szabaly,
                    "leiras": f"{jelolt_nev} tudná betölteni {duty} ügyeletet {day_nap}-án, "
                              f"de ehhez át kellene hágni: {SZABALY_LEIRASOK[szabaly]}.",
                })
            schedule[d][duty] = None
            continue
        candidates.sort(key=lambda x: (x[0], x[1], x[2]))
        chosen = candidates[0][2]
        schedule[d][duty] = chosen
        today_assigned.add(chosen)
        assigned_count[chosen] += 1
        duty_dates[chosen].add(day_date)
        if chosen in RESZ_NAPI_ORASZAMOS:
            kotelezo_ora_used[chosen] += kotelezo_delta_ha_ma_ugyel(chosen, day_date)

    if MIN_SZAKORVOS > 0 and any(schedule[d].values()) and not any(
        tipus_of.get(nm) == "Szakorvos" for nm in schedule[d].values() if nm
    ):
        for duty in ("Stroke", "Aneszt", "Intenzív"):
            current = schedule[d].get(duty)
            if current is None:
                continue
            if (d + 1) in MINDENKEPPEN_SZERETNE.get(current, []):
                continue  # "mindenképp szeretném" alapján kapta ezt a napot - nem cserélhető le
            best = None
            for name, cat, hrs, req, tipus in staff:
                if tipus != "Szakorvos" or not eligible(cat, duty):
                    continue
                if name in today_assigned and name != current:
                    continue
                pref = prefs.get((name, day_date))
                if pref in ("Szabadság", "Nem szeretne"):
                    continue
                if name != current and would_exceed_havi_kvota(name):
                    continue
                if name != current and ugyelet_tiltott(name, day_date):
                    continue
                if name != current and would_exceed_resz_kapacitas(name, day_date):
                    continue
                if name != current and piheno_utkozik(name, day_date):
                    continue
                ratio = assigned_count[name] / target_weight[name]
                if best is None or ratio < best[0]:
                    best = (ratio, name)
            if best is not None and best[1] != current:
                new_name = best[1]
                assigned_count[current] -= 1
                today_assigned.discard(current)
                duty_dates[current].discard(day_date)
                if current in RESZ_NAPI_ORASZAMOS:
                    kotelezo_ora_used[current] -= kotelezo_delta_ha_ma_ugyel(current, day_date)
                schedule[d][duty] = new_name
                today_assigned.add(new_name)
                assigned_count[new_name] += 1
                duty_dates[new_name].add(day_date)
                if new_name in RESZ_NAPI_ORASZAMOS:
                    kotelezo_ora_used[new_name] += kotelezo_delta_ha_ma_ugyel(new_name, day_date)
                break

# ---------------------------------------------------------------------------
# lelépő nap: bármelyik duty utáni nap MINDENKINÉL (napi és havi keretesnél is)
# ---------------------------------------------------------------------------
worked_days = {name: set() for name, *_ in staff}
for d in range(num_days):
    day_date = first_day + datetime.timedelta(days=d)
    for nm in schedule[d].values():
        if nm:
            worked_days[nm].add(day_date)

_ismeretlen_lelepok = ELOZO_HONAP_LELEPOK - set(staff_order_all)
if _ismeretlen_lelepok:
    print(f"FIGYELEM: 'elozo_honap_lelepok'-ban ismeretlen név(ek) - nem lesz hatásuk: {_ismeretlen_lelepok}")

def is_lelepo(name, day_date):
    # Akinek fix, korlátozott heti munkanapja van (pl. Korompai: csak kedd-csütörtök),
    # annál a saját munkanapján soha nem lehet lelépő - az egyetlen lehetséges
    # jelenléti napja nem eshet ki emiatt, se a hónapváltás miatti felülírásból,
    # se a rendes (előző napi ügyeletből eredő) lelépő-logikából.
    if name in RENDES_NAP_CSAK_HETENTE and day_date.weekday() in RENDES_NAP_CSAK_HETENTE[name]:
        return False
    if (day_date - first_day).days == 0 and name in ELOZO_HONAP_LELEPOK:
        return True
    return (day_date - datetime.timedelta(days=1)) in worked_days[name]

# ---------------------------------------------------------------------------
# O1/O2 osztályos kiosztás
# ---------------------------------------------------------------------------
o_assigned_count = {name: 0 for name, *_ in staff}
o1_o2 = {}

def is_available_for_O(name, day_date, exclude=()):
    if name in T_KATEGORIA_NEVEK:
        return False  # T kategória sosem osztályos
    if name in exclude:
        return False
    if name != O1_ALAP and name in schedule[(day_date - first_day).days].values():
        return False  # az alapértelmezett O1-es (pl. Kelemen) akkor is O1 lehet, ha aznap
                       # ügyeletben van - ilyenkor "I/O1", "A/O1" vagy "St/O1" kombinált kód lesz
    if is_szabadsag(name, day_date):
        return False
    if is_lelepo(name, day_date):
        return False
    if heti_fix_esemeny_ma(name, day_date):
        return False
    if name in HAVI_KERETESEK:
        return False
    if name in O1_O2_TILTOTT:
        return False
    if prefs.get((name, day_date)) == "Nem szeretne" and name in CSAK_JELOLT_NAPOKON:
        return False
    if jelenlet_tiltott(name, day_date):
        return False
    if nem_dolgozik_hetente_ma(name, day_date):
        return False
    if name in RESZ_NAPI_ORASZAMOS and not kivansagok[name]["szeret"]:
        return False
    if name in RENDES_NAP_CSAK_HETENTE and day_date.weekday() not in RENDES_NAP_CSAK_HETENTE[name]:
        return False
    if name in KULSOS_GYAKORLATON:
        return False
    return True

def is_available_for_O2(name, day_date, exclude=()):
    """Az O2-nél az St és A ügyeletesek is szóba jöhetnek (nappal még dolgoznak,
    az ügyeletük csak este kezdődik) - csak az Intenzív-ügyeletes van kizárva."""
    if name in T_KATEGORIA_NEVEK:
        return False  # T kategória sosem osztályos
    if name in exclude:
        return False
    if schedule[(day_date - first_day).days].get("Intenzív") == name:
        return False
    if is_szabadsag(name, day_date):
        return False
    if is_lelepo(name, day_date):
        return False
    if heti_fix_esemeny_ma(name, day_date):
        return False
    if name in HAVI_KERETESEK:
        return False
    if name in O1_O2_TILTOTT:
        return False
    if prefs.get((name, day_date)) == "Nem szeretne" and name in CSAK_JELOLT_NAPOKON:
        return False
    if jelenlet_tiltott(name, day_date):
        return False
    if nem_dolgozik_hetente_ma(name, day_date):
        return False
    if name in RESZ_NAPI_ORASZAMOS and not kivansagok[name]["szeret"]:
        return False
    if name in RENDES_NAP_CSAK_HETENTE and day_date.weekday() not in RENDES_NAP_CSAK_HETENTE[name]:
        return False
    if name in KULSOS_GYAKORLATON:
        return False
    return True

def find_substitute(d, day_date, exclude):
    if d + 1 < num_days:
        next_ito = schedule[d + 1].get("Intenzív")
        if next_ito and is_available_for_O(next_ito, day_date, exclude):
            return next_ito
    candidates = [(o_assigned_count[nm], nm) for nm in staff_order_all
                  if is_available_for_O(nm, day_date, exclude)]
    if candidates:
        candidates.sort()
        return candidates[0][1]
    return None

def find_substitute_o2(day_date, exclude):
    d = (day_date - first_day).days
    on_duty_today = set(schedule[d].values())
    # elsőként a nem-ügyeletes (sima "m") jelöltek közül választunk - az St/A ügyeletes
    # csak akkor jön szóba, ha egyáltalán nincs más elérhető, VAGY ha ő a heti
    # folytonosság megtartásához kell (azt a hívó fél külön kezeli, mielőtt idejutna).
    plain_candidates = [(o_assigned_count[nm], nm) for nm in staff_order_all
                         if nm not in on_duty_today and is_available_for_O2(nm, day_date, exclude)]
    if plain_candidates:
        plain_candidates.sort()
        return plain_candidates[0][1]
    duty_candidates = [(o_assigned_count[nm], nm) for nm in staff_order_all
                        if is_available_for_O2(nm, day_date, exclude)]
    if duty_candidates:
        duty_candidates.sort()
        return duty_candidates[0][1]
    return None

def pick_weekly_o2_anchor(d_start):
    """A hét hátralévő hétköznapjaira megnézi, ki elérhető a legtöbb napon O2-re -
    ő lesz a heti "anchor", hogy lehetőleg végig ő maradjon, ne töredezzen szét a hét."""
    het_szam = (first_day + datetime.timedelta(days=d_start)).isocalendar()[1]
    het_napok = [dd for dd in range(d_start, num_days)
                 if (first_day + datetime.timedelta(days=dd)).isocalendar()[1] == het_szam
                 and (first_day + datetime.timedelta(days=dd)).weekday() < 5]
    scores = []
    for nm in staff_order_all:
        if nm in O1_O2_TILTOTT or nm in HAVI_KERETESEK:
            continue
        count = sum(1 for dd in het_napok
                    if is_available_for_O2(nm, first_day + datetime.timedelta(days=dd), exclude=set()))
        if count > 0:
            scores.append((-count, o_assigned_count[nm], nm))
    if scores:
        scores.sort()
        return scores[0][2]
    return None

heti_o2_szemely = None
heti_o2_hetszam = None
heti_anchor_by_het = {}

for d in range(num_days):
    day_date = first_day + datetime.timedelta(days=d)
    if day_date.weekday() >= 5 or not ALT["o1_o2_szukseges_hetkoznap"]:
        continue

    o1 = O1_ALAP if is_available_for_O(O1_ALAP, day_date) else find_substitute(d, day_date, exclude=())
    chosen = {o1} if o1 else set()

    present_count = 0
    ito_present = 1 if schedule[d].get("Intenzív") else 0
    mr_present = 1 if any(heti_fix_esemeny_ma(nm, day_date) for nm in staff_order_all) else 0
    for name in staff_order_all:
        if name in T_KATEGORIA_NEVEK:
            continue  # T kategória sosem számít a műtői/O1-O2 előszámláló jelenlétbe
        if jelenlet_tiltott(name, day_date) and name not in schedule[d].values():
            continue
        if nem_dolgozik_hetente_ma(name, day_date) and name not in schedule[d].values():
            continue
        if (name in RENDES_NAP_CSAK_HETENTE and day_date.weekday() not in RENDES_NAP_CSAK_HETENTE[name]
                and name not in schedule[d].values()):
            continue
        if (name in RESZ_NAPI_ORASZAMOS and not kivansagok[name]["szeret"]
                and name not in schedule[d].values()):
            continue
        if (name in schedule[d].values() or heti_fix_esemeny_ma(name, day_date) or
                (keret_of(name) == "Napi" and not is_szabadsag(name, day_date) and not is_lelepo(name, day_date)
                 and not (name in CSAK_JELOLT_NAPOKON and prefs.get((name, day_date)) == "Nem szeretne"))):
            present_count += 1

    # heti folytonosság: ugyanaz az O2 személy próbál maradni egy héten belül,
    # csak akkor váltunk, ha új hét kezdődik, vagy ha aznap nem elérhető
    het_szam = day_date.isocalendar()[1]
    if het_szam != heti_o2_hetszam:
        heti_o2_hetszam = het_szam
        heti_o2_szemely = pick_weekly_o2_anchor(d)
        heti_anchor_by_het[het_szam] = heti_o2_szemely

    o2_candidate = None
    if O2_ALAP and is_available_for_O2(O2_ALAP, day_date, exclude=chosen):
        o2_candidate = O2_ALAP
    elif heti_o2_szemely and is_available_for_O2(heti_o2_szemely, day_date, exclude=chosen):
        o2_candidate = heti_o2_szemely
    else:
        o2_candidate = find_substitute_o2(day_date, exclude=chosen)

    o2 = None
    if o2_candidate:
        mar_ugyel = schedule[d].get("Stroke") == o2_candidate or schedule[d].get("Aneszt") == o2_candidate
        letszam_csokkenes = 0 if mar_ugyel else 1  # ha már úgyis St/A ügyeletben van, nem vesz el új főt a jelenléti körből
        or_pool_with_o2 = present_count - ito_present - (1 if o1 else 0) - letszam_csokkenes - mr_present
        if or_pool_with_o2 >= MUTO_PADLO:
            o2 = o2_candidate
            if het_szam == heti_o2_hetszam:
                heti_o2_szemely = o2

    if o1:
        o_assigned_count[o1] += 1
    if o2:
        o_assigned_count[o2] += 1
    o1_o2[d] = {"O1": o1, "O2": o2}

# ---------------------------------------------------------------------------
# sanity check
# ---------------------------------------------------------------------------
violations = []
for d in range(num_days):
    day_date = first_day + datetime.timedelta(days=d)
    names_today = [nm for nm in schedule[d].values() if nm]
    if len(names_today) != len(set(names_today)):
        violations.append(f"double-book on {day_date}")
    for duty, nm in schedule[d].items():
        if nm is None:
            continue
        pref = prefs.get((nm, day_date))
        if pref in ("Szabadság", "Nem szeretne"):
            violations.append(f"PREF VIOLATION {nm} {day_date} {pref} -> {duty}")
for name, *_ in staff:
    days_for_name = sorted(first_day + datetime.timedelta(days=d) for d in range(num_days)
                            for duty, nm in schedule[d].items() if nm == name)
    for i in range(1, len(days_for_name)):
        gap = (days_for_name[i] - days_for_name[i - 1]).days
        if gap <= MIN_PIHENO:
            violations.append(f"{name} gap {gap} {days_for_name[i-1]}->{days_for_name[i]}")

print("Napok:", num_days)
print("Violations:", violations if violations else "NONE")
for name, *_ in staff:
    print(f"{name}: {assigned_count[name]} ügyelet")

# ---------------------------------------------------------------------------
# Excel írás
# ---------------------------------------------------------------------------
wb = openpyxl.load_workbook(SABLON_PATH)
ws = wb["Beosztás"]
ws["B3"] = first_day  # a naptár/hét-formulák ehhez igazodnak - a generált hónapra kell állítani
DATA_START = 6
col_map = {"Intenzív": 3, "Stroke": 4, "Aneszt": 5}
for d in range(num_days):
    row = DATA_START + d
    day_date = first_day + datetime.timedelta(days=d)
    for duty, col in col_map.items():
        ws.cell(row=row, column=col, value=schedule[d].get(duty))
    ws.cell(row=row, column=6, value=o1_o2.get(d, {}).get("O1"))
    ws.cell(row=row, column=7, value=o1_o2.get(d, {}).get("O2"))

    # Ellenőrzés oszlop (H) - közvetlenül Python-ban kiszámolva (nem Excel-képlettel), hogy
    # a fájlnak ne kelljen LibreOffice-os újraszámoláson átmennie a kézbesítés előtt.
    uzenetek = []
    intenziv_nm, stroke_nm, aneszt_nm = schedule[d].get("Intenzív"), schedule[d].get("Stroke"), schedule[d].get("Aneszt")
    if day_date.weekday() == 5 and stroke_nm:
        uzenetek.append("Szombaton nincs Stroke ügyelet! ")
    napi_ugyeletesek = [x for x in (intenziv_nm, stroke_nm, aneszt_nm) if x]
    if len(napi_ugyeletesek) != len(set(napi_ugyeletesek)):
        uzenetek.append("Egy fő két ügyeletre nem osztható be egy napon! ")
    for nm in napi_ugyeletesek:
        if any(abs((day_date - dd).days) <= MIN_PIHENO for dd in duty_dates.get(nm, set()) if dd != day_date):
            uzenetek.append("Pihenőidő (min. 2 nap) megsértve! ")
            break
    duty_col_nev = {"Intenzív": intenziv_nm, "Stroke": stroke_nm, "Aneszt": aneszt_nm}
    for duty_nev, nm in duty_col_nev.items():
        if nm and prefs.get((nm, day_date)) == "Szabadság":
            uzenetek.append(f"{nm} szabadságon van ({duty_nev})! ")
    o1_nm, o2_nm = o1_o2.get(d, {}).get("O1"), o1_o2.get(d, {}).get("O2")
    if o1_nm and prefs.get((o1_nm, day_date)) == "Szabadság":
        uzenetek.append(f"{o1_nm} szabadságon van (O1)! ")
    if o2_nm and prefs.get((o2_nm, day_date)) == "Szabadság":
        uzenetek.append(f"{o2_nm} szabadságon van (O2)! ")
    for duty_nev, nm in duty_col_nev.items():
        if nm and prefs.get((nm, day_date)) == "Nem szeretne":
            uzenetek.append(f"{nm} nem szeretett volna dolgozni ({duty_nev})! ")
    if napi_ugyeletesek and not any(tipus_of.get(nm) == "Szakorvos" for nm in napi_ugyeletesek):
        uzenetek.append("Nincs szakorvos az ügyeletben aznap! ")
    if o1_nm and o1_nm == o2_nm:
        uzenetek.append("Az O1 és O2 nem lehet ugyanaz a fő! ")
    if o1_nm and o1_nm != O1_ALAP and o1_nm in napi_ugyeletesek:
        uzenetek.append("Az O1 osztályos nem lehet egyben ügyeletes is! ")
    if o2_nm and o2_nm == intenziv_nm:
        uzenetek.append("Az O2 osztályos nem lehet egyben Intenzív-ügyeletes! ")
    ws.cell(row=row, column=8, value="".join(uzenetek).strip())

# Kívánságok lap feltöltése (a végleges, szabályokkal kiegészített prefs alapján)
ws_kiv = wb["Kívánságok"]
KIV_START = 5
kiv_row_of = {name: KIV_START + i for i, name in enumerate(staff_order_all)}
for (name, day_date), value in prefs.items():
    row = kiv_row_of.get(name)
    if row is None:
        continue
    col = 2 + (day_date - first_day).days
    ws_kiv.cell(row=row, column=col, value=value)

# Dolgozók lap "Kért ügyeletszám" oszlopának feltöltése a havi kívánság-fájlból
ws_staff = wb["Dolgozók"]
STAFF_START = 4
kert = KIV.get("kert_ugyeletszam", {})
NYOLC_ORA_NAPPAL = KIV.get("nyolc_ora_nappal", {})  # nev -> [napok], amikor 8 órás rendes napot tud vállalni
for i, name in enumerate(staff_order_all):
    ws_staff.cell(row=STAFF_START + i, column=4, value=kert.get(name))

ws_print = wb["Nyomtatási beosztás"]
HONAP_NEVEK = ["", "Január", "Február", "Március", "Április", "Május", "Június",
               "Július", "Augusztus", "Szeptember", "Október", "November", "December"]
ws_print["A1"] = f"Nyomtatási beosztás (havi rács) — {HONAP_NEVEK[MONTH]} {YEAR}"
PRINT_HEADER_ROW = 4
PRINT_START = 5
staff_order = staff_order_all
row_of = {name: PRINT_START + i for i, name in enumerate(staff_order)}
muto_row = PRINT_START + len(staff_order)

WEEKEND_FILL = PatternFill("solid", fgColor="C6E0B4")
SZABADSAG_FILL = PatternFill("solid", fgColor="000000")
SZABADSAG_FONT = XLFont(name="Arial", size=9, color="FFFFFF")
NEM_SZERETNE_FILL = PatternFill("solid", fgColor="C23B3B")
NEM_SZERETNE_FONT = XLFont(name="Arial", size=9, color="FFFFFF")
SZERETNE_FILL = PatternFill("solid", fgColor="2E7D4F")
SZERETNE_FONT = XLFont(name="Arial", size=9, color="FFFFFF")
MINDENKEPPEN_FILL = PatternFill("solid", fgColor="C9A227")
MINDENKEPPEN_FONT = XLFont(name="Arial", size=9, color="1C2521")
NYOLCORA_FILL = PatternFill("solid", fgColor="3D6FB4")
NYOLCORA_FONT = XLFont(name="Arial", size=9, color="FFFFFF")

m_count = {name: 0 for name in staff_order}
aktiv_nap_count = {name: 0 for name in staff_order}
hetvegi_ugyelet_count = {name: 0 for name in staff_order}
hetkoznapi_ugyelet_count = {name: 0 for name in staff_order}
lelepo_hetkoznap_count = {name: 0 for name in staff_order}
szabadsag_hetkoznap_count = {name: 0 for name in staff_order}
kulsos_hetkoznap_count = {name: 0 for name in staff_order}  # rezidens külsős gyakorlaton töltött hétköznapja (8 óra/nap jóváírás)
visszahivas_count = {name: 0 for name in REZIDENSEK}
napi_rate_of = {d["nev"]: d["napi_munkaido"] for d in SZAB["dolgozok"]}
raw_present_count_by_day = {}  # d (0-alapú) -> jelenlévők száma aznap, a fő ciklusból

for d in range(num_days):
    day_date = first_day + datetime.timedelta(days=d)
    col = 2 + d
    is_weekend = day_date.weekday() >= 5
    duty_today = {nm: code for code, nm in
                  {"I": schedule[d].get("Intenzív"), "St": schedule[d].get("Stroke"),
                   "A": schedule[d].get("Aneszt")}.items() if nm}
    o_today = o1_o2.get(d, {"O1": None, "O2": None})

    present_count = 0
    ito_present = 1 if schedule[d].get("Intenzív") else 0
    o1_present = 1 if o_today["O1"] else 0
    o2_present = 1 if o_today["O2"] else 0
    mr_present = 0

    if is_weekend:
        ws_print.cell(row=PRINT_HEADER_ROW, column=col).fill = WEEKEND_FILL

    for name in staff_order:
        parts = []
        if name in duty_today:
            parts.append(duty_today[name])
        if name == o_today["O1"]:
            parts.append("O1")
        if name == o_today["O2"]:
            parts.append("O2")
        fix_kod = heti_fix_esemeny_ma(name, day_date)
        if fix_kod and not is_szabadsag(name, day_date):
            parts.append(fix_kod)
            if day_date.weekday() < 5:
                mr_present = 1

        on_szabadsag = is_szabadsag(name, day_date)

        if parts:
            code = "/".join(parts)
        elif is_lelepo(name, day_date):
            code = "el"
        elif on_szabadsag:
            code = ""
        elif is_weekend:
            code = ""
        elif (name in CSAK_JELOLT_NAPOKON and prefs.get((name, day_date)) == "Nem szeretne"
              and (d + 1) not in CSAK_UGYELET_TILTAS_NAPOK.get(name, set())):
            code = ""
        elif name in RENDES_NAP_CSAK_HETENTE and day_date.weekday() not in RENDES_NAP_CSAK_HETENTE[name]:
            code = ""  # fix heti rendes napja van, ezen a hétköznapon nincs bent
        elif jelenlet_tiltott(name, day_date):
            code = ""
        elif nem_dolgozik_hetente_ma(name, day_date):
            code = ""
        elif name in RESZ_NAPI_ORASZAMOS and not kivansagok[name]["szeret"]:
            code = ""  # nincs megadott jó napja - a havi kerete kizárólag ügyeletből teljesül, "m" nap nélkül
        elif name in KULSOS_GYAKORLATON:
            code = ""  # rezidens külsős gyakorlaton - csak visszahívással kaphat "m"-et, ld. lentebb
        elif keret_of(name) != "Napi":
            code = ""
        elif name in T_KATEGORIA_NEVEK:
            code = "t"  # T kategória: minden munkanapon bent van, kis "t" jelöléssel, ügyeletre nem osztható
        else:
            code = "m"

        if code and code != "el":
            aktiv_nap_count[name] += 1
            if name not in T_KATEGORIA_NEVEK:
                present_count += 1
        if name in duty_today:
            if is_weekend:
                hetvegi_ugyelet_count[name] += 1
            else:
                hetkoznapi_ugyelet_count[name] += 1
        if code == "m":
            m_count[name] += 1
        if code == "el" and day_date.weekday() < 5:
            lelepo_hetkoznap_count[name] += 1
        if on_szabadsag and day_date.weekday() < 5:
            szabadsag_hetkoznap_count[name] += 1
        if name in KULSOS_GYAKORLATON and code == "" and day_date.weekday() < 5:
            kulsos_hetkoznap_count[name] += 1
        cell = ws_print.cell(row=row_of[name], column=col, value=code if code else None)
        day_nap = d + 1
        if on_szabadsag:
            cell.fill = SZABADSAG_FILL
            cell.font = SZABADSAG_FONT
        elif day_nap in MINDENKEPPEN_SZERETNE.get(name, []):
            cell.fill = MINDENKEPPEN_FILL
            cell.font = MINDENKEPPEN_FONT
        elif day_nap in NYOLC_ORA_NAPPAL.get(name, []):
            # a "8 óra alkalmas" külön dologról szól (rendes napi munka, nem ügyelet),
            # ezért ez megelőzi a "nem szeretné ügyeletet" jelzést - a kettő nem zárja ki
            # egymást, és a kék jelzés hasznosabb infó, mint az általános piros.
            cell.fill = NYOLCORA_FILL
            cell.font = NYOLCORA_FONT
        elif day_nap in EREDETI_KIVANSAGOK.get(name, {}).get("nem", []):
            cell.fill = NEM_SZERETNE_FILL
            cell.font = NEM_SZERETNE_FONT
        elif day_nap in EREDETI_KIVANSAGOK.get(name, {}).get("szeret", []):
            cell.fill = SZERETNE_FILL
            cell.font = SZERETNE_FONT
        elif is_weekend:
            cell.fill = WEEKEND_FILL

    if is_weekend:
        ws_print.cell(row=muto_row, column=col).fill = WEEKEND_FILL
    else:
        muto_val = present_count - ito_present - o1_present - o2_present - mr_present - MUTO_MIN
        ws_print.cell(row=muto_row, column=col, value=muto_val)
        raw_present_count_by_day[d] = present_count - ito_present - o1_present - o2_present - mr_present

# Pótlólagos "m" napok hozzáadása azoknak, akiknél a megadott "jó napok" nem elegek a
# havi kötelező óraszám eléréséhez - a jelölt napjaik előnyt élveznek (azokat a fő ciklus
# már beírta), de ha ez nem elég, bevonunk plusz napokat is, amíg el nem éri a kapacitást
# (vagy el nem fogynak a lehetséges napok). Ha valaki megadta, mely napokon tud 8 órás
# rendes napot vállalni ("nyolc_ora_nappal"), KIZÁRÓLAG azok közül választhatunk - nem
# találomra bármelyik napról, hiszen ő kifejezetten megmondta, mely napok jók neki. Ha ezt
# nem adta meg, a régi, szabad kereséssel dolgozunk. Akinek fix heti mintája van (pl.
# Korompai: csak kedd-csütörtök), azt ez nem érinti - az egy más típusú, szándékosan
# korlátozott szabály.
for name in RESZ_NAPI_ORASZAMOS:
    if name in RENDES_NAP_CSAK_HETENTE:
        continue
    napi_rate = RESZ_NAPI_ORASZAMOS[name]
    kapacitas = RESZ_NAPI_KAPACITAS[name]
    current_nappali = (NAPI_KOTELEZO_ORA * (aktiv_nap_count[name] - hetvegi_ugyelet_count[name]) +
                        napi_rate * lelepo_hetkoznap_count[name] +
                        napi_rate * szabadsag_hetkoznap_count[name])
    if current_nappali >= kapacitas:
        continue
    megadott_8ora_napok = NYOLC_ORA_NAPPAL.get(name)
    if megadott_8ora_napok is not None:
        jelolt_napok = [d for d in range(num_days) if (d + 1) in megadott_8ora_napok]
    else:
        jelolt_napok = list(range(num_days))
    for d in sorted(jelolt_napok, key=lambda x: raw_present_count_by_day.get(x, 999)):
        if current_nappali >= kapacitas:
            break
        day_date = first_day + datetime.timedelta(days=d)
        if day_date.weekday() >= 5:
            continue
        day_nap = d + 1
        if day_nap in EREDETI_KIFEJEZETT_NEM.get(name, set()):
            continue
        if is_szabadsag(name, day_date):
            continue
        if jelenlet_tiltott(name, day_date):
            continue
        col = 2 + d
        cell = ws_print.cell(row=row_of[name], column=col)
        if cell.value:
            continue
        cell.value = "m"
        m_count[name] += 1
        aktiv_nap_count[name] += 1
        current_nappali += NAPI_KOTELEZO_ORA
        raw_present_count_by_day[d] = raw_present_count_by_day.get(d, 0) + 1
        muto_cell = ws_print.cell(row=muto_row, column=col)
        if muto_cell.value is not None:
            muto_cell.value += 1

# Ugyanez a pótlólagos "8 órás nappali nap" logika a havi keretes (folyamatos munkarendű,
# pl. Pintér Enikő) dolgozóknál is - náluk a "kapacitás" a havi órakeret és a már kapott
# ügyeleti órák különbsége (hiszen egybe számít minden ledolgozott óra).
for name in HAVI_KERETESEK:
    if not havi_oraszam_map.get(name):
        continue
    kvota = havi_oraszam_map[name]
    fennmaradt = kvota - 24 * assigned_count.get(name, 0)
    if fennmaradt <= 0:
        continue
    hozzaadott_nappali_ora = 0
    megadott_8ora_napok = NYOLC_ORA_NAPPAL.get(name)
    if megadott_8ora_napok is not None:
        jelolt_napok = [d for d in range(num_days) if (d + 1) in megadott_8ora_napok]
    else:
        jelolt_napok = list(range(num_days))
    for d in sorted(jelolt_napok, key=lambda x: raw_present_count_by_day.get(x, 999)):
        if hozzaadott_nappali_ora >= fennmaradt:
            break
        day_date = first_day + datetime.timedelta(days=d)
        if day_date.weekday() >= 5:
            continue
        day_nap = d + 1
        if day_nap in EREDETI_KIFEJEZETT_NEM.get(name, set()):
            continue
        if is_szabadsag(name, day_date):
            continue
        if jelenlet_tiltott(name, day_date):
            continue
        col = 2 + d
        cell = ws_print.cell(row=row_of[name], column=col)
        if cell.value:
            continue
        cell.value = "m"
        m_count[name] += 1
        aktiv_nap_count[name] += 1
        hozzaadott_nappali_ora += NAPI_KOTELEZO_ORA
        raw_present_count_by_day[d] = raw_present_count_by_day.get(d, 0) + 1
        muto_cell = ws_print.cell(row=muto_row, column=col)
        if muto_cell.value is not None:
            muto_cell.value += 1

# Túllépés-vágás: ha valakinek (jellemzően nagyon alacsony napi órakeretű, sok "jó napot"
# jelölő résmunkaidős kollégának) a jelölt napjai önmagukban messze meghaladják a kapacitását
# + a 7 órás tűréshatárt, a fölösleges "m" napokat vissza kell venni - ugyanaz a max. 7 órás
# túllépési szabály vonatkozik erre is, mint az ügyeletekre. A legkevésbé műtő-kritikus
# (legtöbb egyéb jelenléttel rendelkező) napokat vesszük vissza először.
for name in RESZ_NAPI_ORASZAMOS:
    if name in RENDES_NAP_CSAK_HETENTE:
        continue
    napi_rate = RESZ_NAPI_ORASZAMOS[name]
    kapacitas = RESZ_NAPI_KAPACITAS[name]
    current_nappali = (NAPI_KOTELEZO_ORA * (aktiv_nap_count[name] - hetvegi_ugyelet_count[name]) +
                        napi_rate * lelepo_hetkoznap_count[name] +
                        napi_rate * szabadsag_hetkoznap_count[name])
    if current_nappali <= kapacitas + TULLEPES_TURESHATAR:
        continue
    m_napok = [d for d in range(num_days)
               if ws_print.cell(row=row_of[name], column=2 + d).value == "m"]
    m_napok.sort(key=lambda x: -raw_present_count_by_day.get(x, 0))
    for d in m_napok:
        if current_nappali <= kapacitas + TULLEPES_TURESHATAR:
            break
        col = 2 + d
        ws_print.cell(row=row_of[name], column=col).value = None
        m_count[name] -= 1
        aktiv_nap_count[name] -= 1
        current_nappali -= NAPI_KOTELEZO_ORA
        raw_present_count_by_day[d] = raw_present_count_by_day.get(d, 1) - 1
        muto_cell = ws_print.cell(row=muto_row, column=col)
        if muto_cell.value is not None:
            muto_cell.value -= 1

# Rezidens-visszahívás: aki külsős gyakorlaton van, azt csak akkor hívjuk vissza "m"-nek,
# ha a műtői minimum (6 fő) máshogy nem teljesülne - méltányosan elosztva a visszahívásokat
# a külsős rezidensek között (mindig a legkevesebbszer visszahívottat választva).
for d in range(num_days):
    day_date = first_day + datetime.timedelta(days=d)
    if day_date.weekday() >= 5:
        continue
    hiany = MUTO_PADLO - raw_present_count_by_day.get(d, 0)
    if hiany <= 0:
        continue
    col = 2 + d
    jeloltek = [nm for nm in REZIDENSEK if nm in KULSOS_GYAKORLATON]
    jeloltek = [nm for nm in jeloltek if ws_print.cell(row=row_of[nm], column=col).value is None
                and not is_szabadsag(nm, day_date) and not is_lelepo(nm, day_date)
                and not jelenlet_tiltott(nm, day_date)]
    jeloltek.sort(key=lambda nm: visszahivas_count[nm])
    for nm in jeloltek:
        if hiany <= 0:
            break
        cell = ws_print.cell(row=row_of[nm], column=col)
        cell.value = "m"
        m_count[nm] += 1
        aktiv_nap_count[nm] += 1
        visszahivas_count[nm] += 1
        kulsos_hetkoznap_count[nm] -= 1  # ezen a napon már nem "hiányzó", hanem "visszahívott"
        raw_present_count_by_day[d] = raw_present_count_by_day.get(d, 0) + 1
        muto_cell = ws_print.cell(row=muto_row, column=col)
        if muto_cell.value is not None:
            muto_cell.value += 1
        hiany -= 1

# Utólagos O2-pótlás: törekedni kell arra, hogy lehetőleg mindennap két osztályos (O1 és
# O2) is legyen, nem csak egy - ezt a fenti pótlólagos "m" feltöltés UTÁN kell megnézni,
# mert lehet, hogy csak az odakerült extra emberek miatt lesz elég a jelenléti létszám.
for d in range(num_days):
    day_date = first_day + datetime.timedelta(days=d)
    if day_date.weekday() >= 5:
        continue
    if o1_o2.get(d, {}).get("O2"):
        continue  # már van O2
    o1_today = o1_o2.get(d, {}).get("O1")
    pool_if_o2 = raw_present_count_by_day.get(d, 0) - 1
    if pool_if_o2 < MUTO_PADLO:
        continue
    exclude = {o1_today} if o1_today else set()
    het_szam_ma = day_date.isocalendar()[1]
    heti_anchor_ma = heti_anchor_by_het.get(het_szam_ma)
    if O2_ALAP and is_available_for_O2(O2_ALAP, day_date, exclude=exclude):
        candidate = O2_ALAP
    elif heti_anchor_ma and is_available_for_O2(heti_anchor_ma, day_date, exclude=exclude):
        candidate = heti_anchor_ma
    else:
        candidate = find_substitute_o2(day_date, exclude=exclude)
    if not candidate:
        continue
    o1_o2.setdefault(d, {"O1": o1_today, "O2": None})["O2"] = candidate
    o_assigned_count[candidate] += 1
    col = 2 + d
    cell = ws_print.cell(row=row_of[candidate], column=col)
    was_blank = cell.value in (None, "")
    if cell.value in (None, "", "m"):
        cell.value = "O2"
    else:
        cell.value = f"{cell.value}/O2"
    if was_blank:
        aktiv_nap_count[candidate] += 1
    raw_present_count_by_day[d] = raw_present_count_by_day.get(d, 0) - 1
    muto_cell = ws_print.cell(row=muto_row, column=col)
    if muto_cell.value is not None:
        muto_cell.value -= 1

# Gulya-prioritás utólagos érvényesítése: Gulya Réka (O2 alapértelmezett) mindig elsőként
# választandó O2-re, ha aznap jelen van - de a "8 órában alkalmas" napok pótlása (ami az
# O1/O2 kiosztás UTÁN fut) miatt előfordulhat, hogy csak most derül ki, hogy aznap jelen
# lesz. Ha valaki más lett O2 helyette, most átvesszük tőle a szerepet.
if O2_ALAP:
    for d in range(num_days):
        day_date = first_day + datetime.timedelta(days=d)
        if day_date.weekday() >= 5:
            continue
        jelenlegi_o2 = o1_o2.get(d, {}).get("O2")
        if jelenlegi_o2 == O2_ALAP:
            continue  # már ő az O2
        col_gulya = 2 + d
        gulya_cell = ws_print.cell(row=row_of[O2_ALAP], column=col_gulya)
        gulya_kod = gulya_cell.value or ""
        gulya_jelen = bool(gulya_kod) and gulya_kod != "el"
        gulya_intenziv = schedule[d].get("Intenzív") == O2_ALAP
        if not gulya_jelen or gulya_intenziv:
            continue
        o1_today = o1_o2.get(d, {}).get("O1")
        if O2_ALAP == o1_today:
            continue  # ő már O1 aznap, nem lehet egyszerre O2 is
        # a jelenlegi O2-est visszaállítjuk: ha volt más szerepe is (pl. "St/O2"), azt megtartja;
        # ha kizárólag O2-ként volt jelen, akkor a rendes "m" jelenlétét kapja vissza (nem tűnhet
        # el a beosztásból csak azért, mert Gulya átvette az O2 szerepet)
        if jelenlegi_o2:
            col_regi = col_gulya
            regi_cell = ws_print.cell(row=row_of[jelenlegi_o2], column=col_regi)
            regi_kod = regi_cell.value or ""
            reszek = regi_kod.split("/")
            reszek = [r for r in reszek if r != "O2"]
            if reszek:
                uj_regi_kod = "/".join(reszek)
            elif keret_of(jelenlegi_o2) == "Napi":
                uj_regi_kod = "m"  # visszakapja a rendes jelenlétét, nem tűnik el
            else:
                uj_regi_kod = None
            regi_cell.value = uj_regi_kod
            o_assigned_count[jelenlegi_o2] -= 1
        # Gulyát beírjuk O2-nek (a meglévő kódjához fűzve, ha van - pl. "m" -> "O2", vagy "St" -> "St/O2")
        if gulya_kod in ("", "m"):
            gulya_cell.value = "O2"
        else:
            gulya_cell.value = f"{gulya_kod}/O2"
        o1_o2.setdefault(d, {"O1": o1_today, "O2": None})["O2"] = O2_ALAP
        o_assigned_count[O2_ALAP] += 1

# Ha egy napon a műtői jelenlét csak egy osztályossal (O1 VAGY O2, de nem mindkettő) jön ki,
# az önmagában -1-gyel rontja a Műtő-sor kijelzett értékét - büntetve, hogy nem sikerült
# mindkét osztályost beosztani, függetlenül a nyers létszámtól.
for d in range(num_days):
    day_date = first_day + datetime.timedelta(days=d)
    if day_date.weekday() >= 5:
        continue
    o_ma = o1_o2.get(d, {"O1": None, "O2": None})
    csak_egy_osztalyos = (1 if o_ma.get("O1") else 0) + (1 if o_ma.get("O2") else 0) == 1
    if csak_egy_osztalyos:
        col = 2 + d
        muto_cell = ws_print.cell(row=muto_row, column=col)
        if muto_cell.value is not None:
            muto_cell.value -= 1

for name in staff_order:
    duty_count = assigned_count.get(name, 0)
    napi_rate = napi_rate_of.get(name, NAPI_KOTELEZO_ORA)
    if keret_of(name) == "Napi":
        # hétvégi ügyelet: a teljes 24 óra ügyeletibe megy, nincs külön 8 órás nappali rész
        nappali_aktiv_napok = aktiv_nap_count[name] - hetvegi_ugyelet_count[name]
        nappali_total = (NAPI_KOTELEZO_ORA * nappali_aktiv_napok +
                          napi_rate * lelepo_hetkoznap_count[name] +
                          napi_rate * szabadsag_hetkoznap_count[name] +
                          napi_rate * kulsos_hetkoznap_count.get(name, 0))
        ugyeleti_total = 16 * hetkoznapi_ugyelet_count[name] + 24 * hetvegi_ugyelet_count[name]
        ws_print.cell(row=row_of[name], column=33, value=nappali_total)
        # Túlóra (nappali): a teljesített nappali óraszám eltérése a szerződéses havi
        # kapacitástól - pozitív, ha túllépte, negatív, ha nem teljesítette. A rész-munkaidősöknél
        # a már meglévő kapacitás-értéket használjuk, teljes állásúaknál ugyanazzal a képlettel
        # (napi óradíj × havi munkanapok) számolva. Kerekítés a matematikai szabály szerint
        # (0,5-től felfelé), egész órára. Berkes Tíbornál nincs túlóra-számítás - ő csak a
        # jelölt napokon jön, nincs havi elvárt kapacitása.
        if name != "Berkes Tíbor":
            kapacitas_altalanos = RESZ_NAPI_KAPACITAS.get(
                name, napi_rate * munkanapok_a_honapban - ELOZO_HONAP_TULORA.get(name, 0))
            tulora = math.floor(nappali_total - kapacitas_altalanos + 0.5)
            ws_print.cell(row=row_of[name], column=35, value=tulora)
    else:
        ugyeleti_total = 24 * duty_count
        if name in HAVI_KERETESEK and havi_oraszam_map.get(name):
            # Folyamatos munkarendű, havi keretes dolgozóknál (pl. Pintér Enikő) a fix
            # napjaik részben 8 órás nappali, részben ügyeleti napok - mindkettő egyben
            # számít a havi órakerethez képest.
            nappali_resz = NAPI_KOTELEZO_ORA * m_count[name]
            ws_print.cell(row=row_of[name], column=33, value=nappali_resz if nappali_resz else None)
            tulora = math.floor(nappali_resz + ugyeleti_total - havi_oraszam_map[name] + 0.5)
            ws_print.cell(row=row_of[name], column=35, value=tulora)
    ws_print.cell(row=row_of[name], column=34, value=ugyeleti_total)

    # Óraelszámolás lap szinkronban tartása ugyanezzel a (helyes, teljes) számítással
    ws_hr = wb["Óraelszámolás"]
    hr_row = row_of[name]  # ugyanaz a sor-illeszkedés, mint a Dolgozók/Nyomtatási beosztás lapokon
    if keret_of(name) == "Napi":
        ws_hr.cell(row=hr_row, column=4, value=nappali_total)
        ws_hr.cell(row=hr_row, column=5, value=ugyeleti_total)
    else:
        ws_hr.cell(row=hr_row, column=6, value=ugyeleti_total)

wb.save(KIMENET_PATH)
print("saved", KIMENET_PATH, f"(seed={SEED})")

if JAVASOLT_KIVETELEK:
    kivetel_kimenet = KIMENET_PATH.rsplit(".", 1)[0] + "_javasolt_kivetelek.json"
    with open(kivetel_kimenet, "w", encoding="utf-8") as f:
        json.dump(JAVASOLT_KIVETELEK, f, ensure_ascii=False, indent=2)
    print(f"Javasolt kivételek ({len(JAVASOLT_KIVETELEK)} db) elmentve: {kivetel_kimenet}")
    for kiv_item in JAVASOLT_KIVETELEK:
        print(" -", kiv_item["leiras"])
