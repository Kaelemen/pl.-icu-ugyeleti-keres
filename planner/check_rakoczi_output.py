import openpyxl
wb = openpyxl.load_workbook('ICU_ugyeleti_beosztas_kimenet.xlsx', data_only=True)
wp = wb['Nyomtatási beosztás']
row_r = None
for r in range(5, 27):
    if wp.cell(row=r, column=1).value == 'Rákóczi Réka':
        row_r = r
napok = [(c - 1, wp.cell(row=row_r, column=c).value) for c in range(2, 32) if wp.cell(row=row_r, column=c).value]
print('Rakoczi minden napja:', napok)
