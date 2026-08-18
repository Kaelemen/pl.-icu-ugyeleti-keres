# -*- coding: utf-8 -*-
import sys
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.protection import SheetProtection

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
KIMENET_PATH = sys.argv[1] if len(sys.argv) > 1 else os.path.join(SCRIPT_DIR, "ICU_ugyeleti_beosztas.xlsx")
KIVANSAGOK_PATH = sys.argv[2] if len(sys.argv) > 2 else None

FONT_NAME = "Arial"

# ---------------------------------------------------------------------------
# Dolgozói törzsadat és szabályok betöltése a szabalyok.json-ból (egyetlen forrás,
# hogy a sablon és a generate_schedule.py sose térjen el egymástól).
# ---------------------------------------------------------------------------
import json as _json
with open(os.path.join(SCRIPT_DIR, "szabalyok.json"), encoding="utf-8") as _f:
    _SZAB = _json.load(_f)

# Ha kaptunk kívánság-fájlt és abban van felülíró "dolgozok" lista (a webes admin
# felület Dolgozók kezelése szekciójából), azzal építjük fel a sablont - így a
# Kívánságok/Ellenőrzés lapok sor-hivatkozásai mindig a TÉNYLEGES, aktuális
# dolgozói listával lesznek szinkronban, sose a generate_schedule.py-val eltérő
# régi adattal.
if KIVANSAGOK_PATH and os.path.exists(KIVANSAGOK_PATH):
    with open(KIVANSAGOK_PATH, encoding="utf-8") as _f:
        _KIV = _json.load(_f)
    if _KIV.get("dolgozok"):
        _SZAB["dolgozok"] = _KIV["dolgozok"]
        print(f"[build.py] Dolgozói törzsadat felülírva a kívánság-fájlból ({len(_KIV['dolgozok'])} fő).")

staff = [(d["nev"], d["kategoria"], d["napi_munkaido"] or 0) for d in _SZAB["dolgozok"]]
O1_ALAP = _SZAB.get("altalanos_szabalyok", {}).get("o1_alapertelmezett_szemely", "")

HAVI_KERETESEK = {d["nev"] for d in _SZAB["dolgozok"]
                   if d["szerzodes_tipus"] == "Részmunkaidő - havi órakeret"}
HAVI_ORASZAM = {d["nev"]: d["havi_oraszam"] for d in _SZAB["dolgozok"] if d.get("havi_oraszam")}

def keret_tipus(name):
    return "Havi" if name in HAVI_KERETESEK else "Napi"

def beosztas_tipus(name):
    for d in _SZAB["dolgozok"]:
        if d["nev"] == name:
            return d["tipus"]
    return "Szakorvos"

MAX_STAFF_ROWS = 60   # buffer rows on Dolgozók / Segéd sheets for future additions
LAST_STAFF_ROW = 1 + MAX_STAFF_ROWS  # row 2 .. LAST_STAFF_ROW

# Kért ügyeletszám: havonta változó adat, a template-ben üresen indul - a
# generate_schedule.py tölti fel a mindenkori kivansagok_ÉÉÉÉ_HH.json alapján.
requested_duties = {}

_OK = _SZAB["orakeret_konstansok"]
NAPI_KOTELEZO_ORA = _OK["napi_kotelezo_ora"]
NAPI_UGYELETI_ORA = _OK["napi_ugyeleti_ora"]
NAPI_LELEPO_ORA = _OK["napi_lelepo_kotelezo_ora"]
HAVI_TELJES_ORA = _OK["havi_teljes_ugyelet_ora"]

wb = Workbook()

# ----------------------------------------------------------------------
# Styles
# ----------------------------------------------------------------------
title_font = Font(name=FONT_NAME, size=16, bold=True, color="1F4E78")
header_font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E78")
input_fill = PatternFill("solid", fgColor="FFFDE7")   # pale yellow = input cell
input_font = Font(name=FONT_NAME, size=11, color="0000FF")  # blue = hardcoded input
normal_font = Font(name=FONT_NAME, size=11)
bold_font = Font(name=FONT_NAME, size=11, bold=True)
thin = Side(style="thin", color="B7B7B7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

WEEKEND_FILL = PatternFill("solid", fgColor="F2F2F2")
ERROR_FILL = PatternFill("solid", fgColor="FFC7CE")
ERROR_FONT = Font(name=FONT_NAME, size=10, bold=True, color="9C0006")

# ========================================================================
# SHEET 1: Súgó (instructions)
# ========================================================================
ws_help = wb.active
ws_help.title = "Súgó"
ws_help.sheet_view.showGridLines = False
ws_help.column_dimensions["A"].width = 100

lines = [
    ("ICU Ügyeleti Beosztás - Használati útmutató", title_font),
    ("", normal_font),
    ("Munkalapok:", bold_font),
    ("  •  Dolgozók - a csapat névsora és jogosultsági kategóriája (I / A / S).", normal_font),
    ("  •  Beosztás - a havi ügyeleti tervező tábla.", normal_font),
    ("  •  Segéd - háttér-számítások, ne módosítsd.", normal_font),
    ("", normal_font),
    ("Jogosultsági kategóriák (Dolgozók lap, 'Kategória' oszlop):", bold_font),
    ("  •  S = csak Stroke ügyeletre osztható be", normal_font),
    ("  •  A = Aneszt és Stroke ügyeletre osztható be", normal_font),
    ("  •  I = Intenzív (ITO), Aneszt és Stroke ügyeletre is osztható be", normal_font),
    ("", normal_font),
    ("Beosztás lap használata:", bold_font),
    ("  1.  A sárga cellába (Hónap kezdete) írd be a tervezett hónap első napját (pl. 2026.08.01).", normal_font),
    ("      A naptár és a napnevek automatikusan felépülnek.", normal_font),
    ("  2.  Az Intenzív / Stroke / Aneszt oszlopokban a legördülő listából válaszd ki a nevet.", normal_font),
    ("      A listák csak a jogosult dolgozókat kínálják fel (Stroke esetén mindenki választható).", normal_font),
    ("  3.  Szombaton nincs Stroke ügyelet - ha mégis kitöltöd, az Ellenőrzés oszlop jelzi a hibát.", normal_font),
    ("  4.  Az Ellenőrzés oszlop automatikusan figyeli:", normal_font),
    ("        -  ugyanaz a személy nem kaphat két ügyeletet ugyanazon a napon,", normal_font),
    ("        -  két ügyelet között legalább 2 napnak el kell telnie (1 pihenőnap),", normal_font),
    ("        -  szombaton nincs Stroke ügyelet.", normal_font),
    ("      Hiba esetén a sor pirossal kiemelődik, és a hibaüzenet megjelenik.", normal_font),
    ("  5.  A hónap első napján a pihenőidő-ellenőrzés még nem tud visszanyúlni az előző", normal_font),
    ("      hónapra - ezt kézzel érdemes ellenőrizni.", normal_font),
    ("", normal_font),
    ("Új dolgozó felvétele:", bold_font),
    ("  •  Írd be a nevet, a kategóriát (I/A/S) és a munkaidő %-ot a Dolgozók lap következő", normal_font),
    ("      üres sorába. A legördülő listák automatikusan frissülnek, nincs szükség egyéb módosításra.", normal_font),
    ("", normal_font),
    ("Munkaidőkeret-arányos ügyeletelosztás (Kimutatás lap):", bold_font),
    ("  •  A Dolgozók lapon minden főhöz megadható a napi munkaidő órában (teljes állás", normal_font),
    ("      etalon = 8 óra/nap; részmunkaidő esetén pl. 4 vagy 6 óra/nap).", normal_font),
    ("  •  A Kimutatás lap automatikusan összeszámolja, hogy az adott hónapban eddig hány", normal_font),
    ("      ügyeletet kapott mindenki, és ezt összeveti a napi munkaidejéhez arányosan elvárt", normal_font),
    ("      ügyeletszámmal (az addig kiosztott összes ügyelet és az aktív dolgozók összesített", normal_font),
    ("      napi óraszáma alapján).", normal_font),
    ("  •  A Státusz oszlop jelzi, ha valaki a munkaidejéhez képest túl sok vagy túl kevés", normal_font),
    ("      ügyeletet kapott (±1 ügyelet eltérésig 'Rendben'-nek számít).", normal_font),
    ("  •  A kimutatás automatikusan frissül, ahogy a Beosztás lapot töltöd.", normal_font),
    ("", normal_font),
    ("Kívánságok és szabadságok (Kívánságok lap):", bold_font),
    ("  •  Minden dolgozóhoz, minden napra megjelölhető: 'Szabadság' (aznap nem osztható be),", normal_font),
    ("      'Nem szeretne' (kerülendő nap) vagy 'Szeretne' (preferált nap).", normal_font),
    ("  •  A hónap automatikusan a Beosztás lapon megadott hónaphoz igazodik.", normal_font),
    ("  •  Ha egy szabadságon lévő vagy 'nem szeretne' jelölésű személyt mégis beosztasz a", normal_font),
    ("      Beosztás lapon, az Ellenőrzés oszlop jelzi.", normal_font),
    ("", normal_font),
    ("Óraelszámolás (Óraelszámolás lap):", bold_font),
    ("  •  A Dolgozók lapon mindenkinél megadható a munkaidőkeret típusa: Napi vagy Havi.", normal_font),
    ("  •  Napi keretes dolgozónál egy ügyelet 8 óra kötelező munkaidőt (az ügyelet napjára) és", normal_font),
    ("      16 óra ügyeleti időt jelent; ha az azt követő nap hétköznap (a dolgozó aznap lép le),", normal_font),
    ("      további 8 óra is a kötelező munkaidőbe számít.", normal_font),
    ("  •  Havi keretes dolgozónál a teljes 24 óra egyben számít bele a havi munkaidőkeretbe.", normal_font),
    ("  •  Az Óraelszámolás lap automatikusan összesíti ezeket a Beosztás lap alapján.", normal_font),
    ("", normal_font),
    ("Szakorvosi lefedettség:", bold_font),
    ("  •  A Dolgozók lapon mindenkinél megadható: Szakorvos vagy Rezidens.", normal_font),
    ("  •  Minden napon, amikor van kitöltött ügyelet, legalább 1 szakorvosnak kell lennie a", normal_font),
    ("      beosztott 3 ügyeletes között - ha nincs, az Ellenőrzés oszlop jelzi.", normal_font),
    ("", normal_font),
    ("Osztályos fedezet (O1/O2):", bold_font),
    ("  •  Hétköznaponként (hétfő-péntek) 2 fő osztályos szükséges az ügyeletes mellett -", normal_font),
    ("      ezt az Osztályos 1 és Osztályos 2 oszlopokban jelölheted.", normal_font),
    ("", normal_font),
    ("Színjelölés:", bold_font),
    ("  •  Sárga cella = kézzel kitöltendő mező.", normal_font),
    ("  •  Szürke sor = hétvége (szombat/vasárnap).", normal_font),
    ("  •  Piros sor / szöveg = szabálysértés.", normal_font),
]
for i, (text, font) in enumerate(lines, start=1):
    c = ws_help.cell(row=i, column=1, value=text)
    c.font = font
    c.alignment = left

# ========================================================================
# SHEET 2: Dolgozók
# ========================================================================
ws_staff = wb.create_sheet("Dolgozók")
ws_staff.sheet_view.showGridLines = False
ws_staff.column_dimensions["A"].width = 28
ws_staff.column_dimensions["B"].width = 14
ws_staff.column_dimensions["C"].width = 14
ws_staff.column_dimensions["D"].width = 16
ws_staff.column_dimensions["E"].width = 14
ws_staff.column_dimensions["F"].width = 18
ws_staff.column_dimensions["G"].width = 10
ws_staff.column_dimensions["H"].width = 10
ws_staff.column_dimensions["I"].width = 16

ws_staff["A1"] = "Dolgozók névsora és jogosultsága"
ws_staff["A1"].font = title_font
ws_staff.merge_cells("A1:I1")

ws_staff["A2"] = ("Kategória: S=csak Stroke | A=Aneszt+Stroke | I=Intenzív+Aneszt+Stroke.  "
                   "Napi munkaidő: teljes állás etalon = 8 óra/nap, részmunkaidőnél pl. 4/6 óra/nap.  "
                   "Kért ügyeletszám: opcionális - ha egy dolgozó konkrét számot kért, a Kimutatás lap "
                   "ezt veszi célértéknek az arányos elosztás helyett; üresen hagyva az arányos (óra "
                   "alapú) célérték érvényes.  Munkaidőkeret típusa: Napi = napi 8 órás keret "
                   "(ügyelet: napi 8 óra + lelépő hétköznapon +8 óra kötelezőbe, 16 óra ügyeletibe); "
                   "Havi = a 24 órás ügyelet egyben számít a havi keretbe (ld. Óraelszámolás lap).  "
                   "Szakorvos/Rezidens: minden napi ügyeletben legalább 1 szakorvosnak lennie kell "
                   "(ld. Beosztás lap Ellenőrzés oszlopa).  Havi óraszám: csak Havi keretes dolgozóknál "
                   "kötelező - a hónapban kiosztott ügyeletek (24 óra/ügyelet) ezt nem léphetik túl; "
                   "Havi keretes dolgozó nem kap O1/O2 osztályos beosztást, kizárólag ügyeletet.")
ws_staff["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color="808080")
ws_staff["A2"].alignment = left
ws_staff.merge_cells("A2:I2")
ws_staff.row_dimensions[2].height = 80

headers = ["Név", "Kategória (I/A/S)", "Napi munkaidő (óra)", "Kért ügyeletszám (hó)",
           "Munkaidőkeret típusa", "Szakorvos/Rezidens", "Rang - Intenzív", "Rang - Aneszt",
           "Havi óraszám (csak Havi keretesnek)"]
for j, h in enumerate(headers, start=1):
    c = ws_staff.cell(row=3, column=j, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

row0 = 4
for i, (name, cat, pct) in enumerate(staff):
    r = row0 + i
    c_name = ws_staff.cell(row=r, column=1, value=name)
    c_name.font = input_font
    c_name.fill = input_fill
    c_name.border = border
    c_cat = ws_staff.cell(row=r, column=2, value=cat)
    c_cat.font = input_font
    c_cat.fill = input_fill
    c_cat.alignment = center
    c_cat.border = border
    c_pct = ws_staff.cell(row=r, column=3, value=pct)
    c_pct.font = input_font
    c_pct.fill = input_fill
    c_pct.alignment = center
    c_pct.border = border
    c_pct.number_format = '0.0" óra"'
    req = requested_duties.get(name)
    c_req = ws_staff.cell(row=r, column=4, value=req)
    c_req.font = input_font
    c_req.fill = input_fill
    c_req.alignment = center
    c_req.border = border
    c_ket = ws_staff.cell(row=r, column=5, value=keret_tipus(name))
    c_ket.font = input_font
    c_ket.fill = input_fill
    c_ket.alignment = center
    c_ket.border = border
    c_tip = ws_staff.cell(row=r, column=6, value=beosztas_tipus(name))
    c_tip.font = input_font
    c_tip.fill = input_fill
    c_tip.alignment = center
    c_tip.border = border
    havi_ora = HAVI_ORASZAM.get(name)
    c_havi = ws_staff.cell(row=r, column=9, value=havi_ora)
    c_havi.font = input_font
    c_havi.fill = input_fill
    c_havi.alignment = center
    c_havi.border = border
    c_havi.number_format = '0" óra"'

last_data_row = row0 + MAX_STAFF_ROWS - 1  # leave buffer rows for future staff
for r in range(row0, last_data_row + 1):
    # rank formulas (helper, used by Segéd sheet lookups)
    c_rank_i = ws_staff.cell(
        row=r, column=7,
        value=f'=IF($B{r}="I",COUNTIF($B${row0}:$B{r},"I"),"")'
    )
    c_rank_i.font = Font(name=FONT_NAME, size=10, color="808080")
    c_rank_i.alignment = center
    c_rank_a = ws_staff.cell(
        row=r, column=8,
        value=f'=IF(OR($B{r}="I",$B{r}="A"),COUNTIF($B${row0}:$B{r},"I")+COUNTIF($B${row0}:$B{r},"A"),"")'
    )
    c_rank_a.font = Font(name=FONT_NAME, size=10, color="808080")
    c_rank_a.alignment = center
    ws_staff.cell(row=r, column=1).border = border
    ws_staff.cell(row=r, column=2).border = border
    ws_staff.cell(row=r, column=3).border = border
    ws_staff.cell(row=r, column=4).border = border
    ws_staff.cell(row=r, column=5).border = border
    ws_staff.cell(row=r, column=6).border = border
    ws_staff.cell(row=r, column=9).border = border

# category dropdown (I/A/S) for the whole buffer range
dv_cat = DataValidation(type="list", formula1='"I,A,S"', allow_blank=True, showErrorMessage=True)
dv_cat.error = "Csak I, A vagy S adható meg."
dv_cat.errorTitle = "Érvénytelen kategória"
ws_staff.add_data_validation(dv_cat)
dv_cat.add(f"B{row0}:B{last_data_row}")

# napi munkaidő (óra) validation (0-24)
dv_pct = DataValidation(type="decimal", operator="between", formula1="0", formula2="24",
                         allow_blank=True, showErrorMessage=True)
dv_pct.error = "A napi munkaidő 0 és 24 óra között adható meg."
dv_pct.errorTitle = "Érvénytelen érték"
ws_staff.add_data_validation(dv_pct)
dv_pct.add(f"C{row0}:C{last_data_row}")

# kért ügyeletszám validation (0-31, egész szám)
dv_req = DataValidation(type="whole", operator="between", formula1="0", formula2="31",
                         allow_blank=True, showErrorMessage=True)
dv_req.error = "A kért ügyeletszám 0 és 31 között adható meg."
dv_req.errorTitle = "Érvénytelen érték"
ws_staff.add_data_validation(dv_req)
dv_req.add(f"D{row0}:D{last_data_row}")

# munkaidőkeret típusa (Napi/Havi) dropdown
dv_ket = DataValidation(type="list", formula1='"Napi,Havi"', allow_blank=True, showErrorMessage=True)
dv_ket.error = "Csak Napi vagy Havi adható meg."
dv_ket.errorTitle = "Érvénytelen érték"
ws_staff.add_data_validation(dv_ket)
dv_ket.add(f"E{row0}:E{last_data_row}")

# szakorvos/rezidens dropdown
dv_tip = DataValidation(type="list", formula1='"Szakorvos,Rezidens"', allow_blank=True, showErrorMessage=True)
dv_tip.error = "Csak Szakorvos vagy Rezidens adható meg."
dv_tip.errorTitle = "Érvénytelen érték"
ws_staff.add_data_validation(dv_tip)
dv_tip.add(f"F{row0}:F{last_data_row}")

# havi óraszám validáció (0-200)
dv_havi = DataValidation(type="decimal", operator="between", formula1="0", formula2="200",
                          allow_blank=True, showErrorMessage=True)
dv_havi.error = "A havi óraszám 0 és 200 között adható meg."
dv_havi.errorTitle = "Érvénytelen érték"
ws_staff.add_data_validation(dv_havi)
dv_havi.add(f"I{row0}:I{last_data_row}")

ws_staff.freeze_panes = "A4"

# ========================================================================
# SHEET 3: Segéd (helper lookups - filtered eligibility lists + weekday names)
# ========================================================================
ws_help2 = wb.create_sheet("Segéd")
ws_help2.sheet_view.showGridLines = False
ws_help2.sheet_state = "visible"
ws_help2.column_dimensions["A"].width = 26
ws_help2.column_dimensions["B"].width = 26
ws_help2.column_dimensions["C"].width = 3
ws_help2.column_dimensions["E"].width = 14

ws_help2["A1"] = "Intenzív jogosultak"
ws_help2["B1"] = "Aneszt jogosultak"
for cell in ("A1", "B1"):
    ws_help2[cell].font = header_font
    ws_help2[cell].fill = header_fill
    ws_help2[cell].alignment = center

for i in range(MAX_STAFF_ROWS):
    r = 2 + i
    f_i = (
        f'=IFERROR(INDEX(Dolgozók!$A${row0}:$A${last_data_row},'
        f'MATCH({i+1},Dolgozók!$G${row0}:$G${last_data_row},0)),"")'
    )
    f_a = (
        f'=IFERROR(INDEX(Dolgozók!$A${row0}:$A${last_data_row},'
        f'MATCH({i+1},Dolgozók!$H${row0}:$H${last_data_row},0)),"")'
    )
    ws_help2.cell(row=r, column=1, value=f_i).font = normal_font
    ws_help2.cell(row=r, column=2, value=f_a).font = normal_font

# Hungarian weekday names, ordered to match WEEKDAY(date,2): 1=Monday ... 7=Sunday
ws_help2["E1"] = "Napnevek"
ws_help2["E1"].font = header_font
ws_help2["E1"].fill = header_fill
ws_help2["E1"].alignment = center
weekday_names = ["Hétfő", "Kedd", "Szerda", "Csütörtök", "Péntek", "Szombat", "Vasárnap"]
for i, wd in enumerate(weekday_names):
    ws_help2.cell(row=2 + i, column=5, value=wd).font = normal_font

# ========================================================================
# Named ranges
# ========================================================================
def add_name(name, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=ref)

add_name("IntenzivLista", f"Segéd!$A$2:$A${1+MAX_STAFF_ROWS}")
add_name("AnesztLista", f"Segéd!$B$2:$B${1+MAX_STAFF_ROWS}")
add_name("StrokeLista", f"Dolgozók!$A${row0}:$A${last_data_row}")

# ========================================================================
# SHEET: Kívánságok (staff preferences & vacation grid)
# ========================================================================
ws_kiv = wb.create_sheet("Kívánságok")
ws_kiv.sheet_view.showGridLines = False

ws_kiv.column_dimensions["A"].width = 26
for i in range(31):
    ws_kiv.column_dimensions[get_column_letter(2 + i)].width = 6.5

ws_kiv["A1"] = "Dolgozói kívánságok és szabadságok"
ws_kiv["A1"].font = title_font
ws_kiv.merge_cells("A1:AF1")

ws_kiv["A2"] = ('Jelöld napokra: "Szabadság" (nem osztható be), "Nem szeretne" (kerülendő), '
                 '"Szeretne" (preferált). A hónap a Beosztás lap "Hónap kezdete" mezőjéhez igazodik.')
ws_kiv["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color="808080")
ws_kiv["A2"].alignment = left
ws_kiv.merge_cells("A2:AF2")
ws_kiv.row_dimensions[2].height = 28

KIV_HEADER_ROW = 4
KIV_START = KIV_HEADER_ROW + 1
KIV_END = KIV_START + MAX_STAFF_ROWS - 1

c_name_hdr = ws_kiv.cell(row=KIV_HEADER_ROW, column=1, value="Név")
c_name_hdr.font = header_font
c_name_hdr.fill = header_fill
c_name_hdr.alignment = center
c_name_hdr.border = border

for i in range(31):
    col = 2 + i
    f_hdr = f'=IF({i}<DAY(EOMONTH(Beosztás!$B$3,0)),Beosztás!$B$3+{i},"")'
    c_hdr = ws_kiv.cell(row=KIV_HEADER_ROW, column=col, value=f_hdr)
    c_hdr.number_format = "mm.dd."
    c_hdr.font = header_font
    c_hdr.fill = header_fill
    c_hdr.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)
    c_hdr.border = border

for r in range(row0, last_data_row + 1):
    sr = r + (KIV_START - row0)
    c_name = ws_kiv.cell(row=sr, column=1, value=f'=Dolgozók!$A{r}')
    c_name.font = normal_font
    c_name.border = border
    for i in range(31):
        col = 2 + i
        c = ws_kiv.cell(row=sr, column=col, value=None)
        c.font = Font(name=FONT_NAME, size=9)
        c.alignment = center
        c.border = border

dv_kiv = DataValidation(type="list", formula1='"Szabadság,Nem szeretne,Szeretne"',
                         allow_blank=True, showErrorMessage=True)
dv_kiv.error = 'Csak "Szabadság", "Nem szeretne" vagy "Szeretne" adható meg.'
dv_kiv.errorTitle = "Érvénytelen érték"
ws_kiv.add_data_validation(dv_kiv)
dv_kiv.add(f"B{KIV_START}:AF{KIV_END}")

VACATION_FILL = PatternFill("solid", fgColor="FFC7CE")
AVOID_FILL = PatternFill("solid", fgColor="FFE699")
PREFER_FILL = PatternFill("solid", fgColor="C6E0B4")
kiv_range = f"B{KIV_START}:AF{KIV_END}"
ws_kiv.conditional_formatting.add(kiv_range, FormulaRule(formula=[f'B{KIV_START}="Szabadság"'], fill=VACATION_FILL))
ws_kiv.conditional_formatting.add(kiv_range, FormulaRule(formula=[f'B{KIV_START}="Nem szeretne"'], fill=AVOID_FILL))
ws_kiv.conditional_formatting.add(kiv_range, FormulaRule(formula=[f'B{KIV_START}="Szeretne"'], fill=PREFER_FILL))

ws_kiv.freeze_panes = f"B{KIV_START}"

add_name("KivNevek", f"Kívánságok!$A${KIV_START}:$A${KIV_END}")
add_name("KivAdat", f"Kívánságok!$B${KIV_START}:$AF${KIV_END}")
add_name("KivFejlec", f"Kívánságok!$B${KIV_HEADER_ROW}:$AF${KIV_HEADER_ROW}")

# ========================================================================
# SHEET 4: Beosztás (the actual monthly schedule)
# ========================================================================
ws = wb.create_sheet("Beosztás")
ws.sheet_view.showGridLines = False
wb.move_sheet("Beosztás", offset=-2)  # put it right after Súgó, before Dolgozók
wb.active = 0

ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 20
ws.column_dimensions["F"].width = 18
ws.column_dimensions["G"].width = 18
ws.column_dimensions["H"].width = 60
for col in ("I", "J", "K", "L", "M", "N", "O", "P"):
    ws.column_dimensions[col].width = 14
    ws.column_dimensions[col].hidden = True

ws["A1"] = "ICU Ügyeleti Beosztás"
ws["A1"].font = title_font
ws.merge_cells("A1:H1")

ws["A3"] = "Hónap kezdete:"
ws["A3"].font = bold_font
import datetime
ws["B3"] = datetime.date(2026, 8, 1)
ws["B3"].number_format = "yyyy.mm.dd."
ws["B3"].font = input_font
ws["B3"].fill = input_fill
ws["B3"].alignment = center
ws["B3"].border = border
ws["C3"] = "<- írd be a hónap első napját, a naptár automatikusan frissül"
ws["C3"].font = Font(name=FONT_NAME, size=10, italic=True, color="808080")
ws.merge_cells("C3:H3")

HEADER_ROW = 5
DATA_START = 6
DATA_END = DATA_START + 30  # 31 potential days

headers = ["Dátum", "Nap", "Intenzív", "Stroke", "Aneszt", "Osztályos 1 (O1)",
           "Osztályos 2 (O2)", "Ellenőrzés"]
for j, h in enumerate(headers, start=1):
    c = ws.cell(row=HEADER_ROW, column=j, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

for i in range(31):
    r = DATA_START + i
    # Dátum
    f_date = f'=IF({i}<DAY(EOMONTH($B$3,0)),$B$3+{i},"")'
    c_date = ws.cell(row=r, column=1, value=f_date)
    c_date.number_format = "yyyy.mm.dd."
    c_date.font = normal_font
    c_date.alignment = center
    c_date.border = border

    # Nap (weekday name, locale-independent lookup)
    f_day = f'=IF($A{r}="","",INDEX(Segéd!$E$2:$E$8,WEEKDAY($A{r},2)))'
    c_day = ws.cell(row=r, column=2, value=f_day)
    c_day.font = normal_font
    c_day.alignment = center
    c_day.border = border

    # Input columns: Intenzív, Stroke, Aneszt, Osztályos 1, Osztályos 2
    for col in (3, 4, 5, 6, 7):
        c = ws.cell(row=r, column=col, value=None)
        c.font = input_font
        c.fill = input_fill
        c.alignment = center
        c.border = border

    # Ellenőrzés formula
    # Rest-day check: only the immediately preceding day matters. A gap of exactly
    # 2 days (duty, 1 rest day, duty) satisfies "min. 2 nap" and must NOT be flagged;
    # only a gap of 0-1 days (no rest day) is a violation.
    prev_row = r - 1
    dup_check = (
        f'IF(OR(AND($C{r}<>"",$C{r}=$D{r}),AND($C{r}<>"",$C{r}=$E{r}),'
        f'AND($D{r}<>"",$D{r}=$E{r})),"Egy fő két ügyeletre nem osztható be egy napon! ","")'
    )
    sat_check = (
        f'IF(AND(WEEKDAY($A{r},2)=6,$D{r}<>""),"Szombaton nincs Stroke ügyelet! ","")'
    )
    if prev_row >= DATA_START:
        rest_check = (
            f'IF(SUMPRODUCT(($C{r}:$E{r}<>"")*COUNTIF($C{prev_row}:$E{prev_row},$C{r}:$E{r}))>0,'
            f'"Pihenoido (min. 2 nap) megsertve! ","")'
        ).replace("Pihenoido (min. 2 nap) megsertve!", "Pihenőidő (min. 2 nap) megsértve!")
    else:
        rest_check = '""'

    # Helper lookups: staff preference/vacation on this date, per duty column
    i_pref = (
        f'=IF($C{r}="","",IFERROR(INDEX(KivAdat,MATCH($C{r},KivNevek,0),MATCH($A{r},KivFejlec,0)),""))'
    )
    j_pref = (
        f'=IF($D{r}="","",IFERROR(INDEX(KivAdat,MATCH($D{r},KivNevek,0),MATCH($A{r},KivFejlec,0)),""))'
    )
    k_pref = (
        f'=IF($E{r}="","",IFERROR(INDEX(KivAdat,MATCH($E{r},KivNevek,0),MATCH($A{r},KivFejlec,0)),""))'
    )
    ws.cell(row=r, column=9, value=i_pref).font = Font(name=FONT_NAME, size=9)
    ws.cell(row=r, column=10, value=j_pref).font = Font(name=FONT_NAME, size=9)
    ws.cell(row=r, column=11, value=k_pref).font = Font(name=FONT_NAME, size=9)

    # Helper lookups: szakorvos/rezidens típus, per duty column
    l_tip = (
        f'=IF($C{r}="","",IFERROR(INDEX(Dolgozók!$F${row0}:$F${last_data_row},'
        f'MATCH($C{r},Dolgozók!$A${row0}:$A${last_data_row},0)),""))'
    )
    m_tip = (
        f'=IF($D{r}="","",IFERROR(INDEX(Dolgozók!$F${row0}:$F${last_data_row},'
        f'MATCH($D{r},Dolgozók!$A${row0}:$A${last_data_row},0)),""))'
    )
    n_tip = (
        f'=IF($E{r}="","",IFERROR(INDEX(Dolgozók!$F${row0}:$F${last_data_row},'
        f'MATCH($E{r},Dolgozók!$A${row0}:$A${last_data_row},0)),""))'
    )
    ws.cell(row=r, column=12, value=l_tip).font = Font(name=FONT_NAME, size=9)
    ws.cell(row=r, column=13, value=m_tip).font = Font(name=FONT_NAME, size=9)
    ws.cell(row=r, column=14, value=n_tip).font = Font(name=FONT_NAME, size=9)

    # Helper lookups: preference/vacation for O1/O2
    o_pref = (
        f'=IF($F{r}="","",IFERROR(INDEX(KivAdat,MATCH($F{r},KivNevek,0),MATCH($A{r},KivFejlec,0)),""))'
    )
    p_pref = (
        f'=IF($G{r}="","",IFERROR(INDEX(KivAdat,MATCH($G{r},KivNevek,0),MATCH($A{r},KivFejlec,0)),""))'
    )
    ws.cell(row=r, column=15, value=o_pref).font = Font(name=FONT_NAME, size=9)
    ws.cell(row=r, column=16, value=p_pref).font = Font(name=FONT_NAME, size=9)

    vac_check = (
        f'IF($I{r}="Szabadság",$C{r}&" szabadságon van (Intenzív)! ","")'
        f'&IF($J{r}="Szabadság",$D{r}&" szabadságon van (Stroke)! ","")'
        f'&IF($K{r}="Szabadság",$E{r}&" szabadságon van (Aneszt)! ","")'
        f'&IF($O{r}="Szabadság",$F{r}&" szabadságon van (O1)! ","")'
        f'&IF($P{r}="Szabadság",$G{r}&" szabadságon van (O2)! ","")'
    )
    pref_check = (
        f'IF($I{r}="Nem szeretne",$C{r}&" nem szeretett volna dolgozni (Intenzív)! ","")'
        f'&IF($J{r}="Nem szeretne",$D{r}&" nem szeretett volna dolgozni (Stroke)! ","")'
        f'&IF($K{r}="Nem szeretne",$E{r}&" nem szeretett volna dolgozni (Aneszt)! ","")'
    )
    szak_check = (
        f'IF(AND(OR($C{r}<>"",$D{r}<>"",$E{r}<>""),COUNTIF($L{r}:$N{r},"Szakorvos")=0),'
        f'"Nincs szakorvos az ügyeletben aznap! ","")'
    )
    o_check = (
        f'IF(AND($F{r}<>"",$F{r}=$G{r}),"Az O1 és O2 nem lehet ugyanaz a fő! ","")'
        f'&IF(AND($F{r}<>"",$F{r}<>"{O1_ALAP}",OR($F{r}=$C{r},$F{r}=$D{r},$F{r}=$E{r})),'
        f'"Az O1 osztályos nem lehet egyben ügyeletes is! ","")'
        f'&IF(AND($G{r}<>"",$G{r}=$C{r}),'
        f'"Az O2 osztályos nem lehet egyben Intenzív-ügyeletes! ","")'
    )

    f_err = (
        f'=IF($A{r}="","",TRIM({sat_check}&{dup_check}&{rest_check}&{vac_check}&'
        f'{pref_check}&{szak_check}&{o_check}))'
    )
    c_err = ws.cell(row=r, column=8, value=f_err)
    c_err.font = Font(name=FONT_NAME, size=10)
    c_err.alignment = left
    c_err.border = border

# Data validation dropdowns
dv_int = DataValidation(type="list", formula1="=IntenzivLista", allow_blank=True, showErrorMessage=True)
dv_int.error = "A kiválasztott személy nem jogosult Intenzív ügyeletre."
dv_int.errorTitle = "Nem jogosult"
ws.add_data_validation(dv_int)
dv_int.add(f"C{DATA_START}:C{DATA_END}")

dv_stroke = DataValidation(type="list", formula1="=StrokeLista", allow_blank=True, showErrorMessage=True)
dv_stroke.error = "A kiválasztott személy nem szerepel a névsorban."
dv_stroke.errorTitle = "Nem jogosult"
ws.add_data_validation(dv_stroke)
dv_stroke.add(f"D{DATA_START}:D{DATA_END}")

dv_anesz = DataValidation(type="list", formula1="=AnesztLista", allow_blank=True, showErrorMessage=True)
dv_anesz.error = "A kiválasztott személy nem jogosult Aneszt ügyeletre."
dv_anesz.errorTitle = "Nem jogosult"
ws.add_data_validation(dv_anesz)
dv_anesz.add(f"E{DATA_START}:E{DATA_END}")

dv_o1 = DataValidation(type="list", formula1="=StrokeLista", allow_blank=True, showErrorMessage=True)
dv_o1.error = "A kiválasztott személy nem szerepel a névsorban."
dv_o1.errorTitle = "Nem jogosult"
ws.add_data_validation(dv_o1)
dv_o1.add(f"F{DATA_START}:F{DATA_END}")

dv_o2 = DataValidation(type="list", formula1="=StrokeLista", allow_blank=True, showErrorMessage=True)
dv_o2.error = "A kiválasztott személy nem szerepel a névsorban."
dv_o2.errorTitle = "Nem jogosult"
ws.add_data_validation(dv_o2)
dv_o2.add(f"G{DATA_START}:G{DATA_END}")

# Conditional formatting
err_rule = FormulaRule(formula=[f'$H{DATA_START}<>""'], fill=ERROR_FILL, stopIfTrue=False)
ws.conditional_formatting.add(f"A{DATA_START}:H{DATA_END}", err_rule)

err_font_rule = FormulaRule(formula=[f'$H{DATA_START}<>""'], font=ERROR_FONT)
ws.conditional_formatting.add(f"H{DATA_START}:H{DATA_END}", err_font_rule)

weekend_rule = FormulaRule(
    formula=[f'OR(WEEKDAY($A{DATA_START},2)=6,WEEKDAY($A{DATA_START},2)=7)'],
    fill=WEEKEND_FILL,
    stopIfTrue=False,
)
ws.conditional_formatting.add(f"A{DATA_START}:B{DATA_END}", weekend_rule)

ws.freeze_panes = f"A{DATA_START}"

# ========================================================================
# SHEET 5: Kimutatás (workload-proportional duty tracking)
# ========================================================================
ws_stat = wb.create_sheet("Kimutatás")
ws_stat.sheet_view.showGridLines = False
wb.move_sheet("Kimutatás", offset=-1)  # place right after Dolgozók, before Segéd

ws_stat.column_dimensions["A"].width = 26
ws_stat.column_dimensions["B"].width = 14
ws_stat.column_dimensions["C"].width = 16
ws_stat.column_dimensions["D"].width = 20
ws_stat.column_dimensions["E"].width = 14
ws_stat.column_dimensions["F"].width = 12
ws_stat.column_dimensions["G"].width = 20

ws_stat["A1"] = "Munkaidő-arányos ügyeletelosztás"
ws_stat["A1"].font = title_font
ws_stat.merge_cells("A1:G1")

ws_stat["A2"] = "Összes kiosztott ügyelet eddig:"
ws_stat["A2"].font = bold_font
ws_stat["B2"] = f"=COUNTA(Beosztás!$C${DATA_START}:$E${DATA_END})"
ws_stat["B2"].font = normal_font
ws_stat["B2"].alignment = center

ws_stat["A3"] = "Aktív dolgozók összesített napi munkaideje (óra):"
ws_stat["A3"].font = bold_font
ws_stat["B3"] = f'=SUMIF(Dolgozók!$A${row0}:$A${last_data_row},"<>",Dolgozók!$C${row0}:$C${last_data_row})'
ws_stat["B3"].font = normal_font
ws_stat["B3"].alignment = center

STAT_HEADER_ROW = 5
stat_headers = ["Név", "Napi munkaidő (óra)", "Kért ügyeletszám", "Ügyeletek száma (hó)",
                 "Elvárt arány", "Eltérés", "Státusz"]
for j, h in enumerate(stat_headers, start=1):
    c = ws_stat.cell(row=STAT_HEADER_ROW, column=j, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

stat_row_offset = STAT_HEADER_ROW + 1 - row0  # so stat row aligns 1:1 with Dolgozók row
for r in range(row0, last_data_row + 1):
    sr = r + stat_row_offset
    c_name = ws_stat.cell(row=sr, column=1, value=f'=Dolgozók!$A{r}')
    c_name.font = normal_font
    c_name.border = border

    c_pct = ws_stat.cell(row=sr, column=2, value=f'=IF($A{sr}="","",Dolgozók!$C{r})')
    c_pct.font = normal_font
    c_pct.alignment = center
    c_pct.number_format = '0.0" óra"'
    c_pct.border = border

    c_req = ws_stat.cell(row=sr, column=3, value=f'=IF($A{sr}="","",Dolgozók!$D{r})')
    c_req.font = normal_font
    c_req.alignment = center
    c_req.border = border

    c_count = ws_stat.cell(
        row=sr, column=4,
        value=f'=IF($A{sr}="","",COUNTIF(Beosztás!$C${DATA_START}:$E${DATA_END},$A{sr}))'
    )
    c_count.font = normal_font
    c_count.alignment = center
    c_count.border = border

    c_target = ws_stat.cell(
        row=sr, column=5,
        value=f'=IF($A{sr}="","",IF($C{sr}<>"",$C{sr},IF($B$3=0,0,$B{sr}/$B$3*$B$2)))'
    )
    c_target.font = normal_font
    c_target.alignment = center
    c_target.number_format = "0.0"
    c_target.border = border

    c_diff = ws_stat.cell(
        row=sr, column=6,
        value=f'=IF($A{sr}="","",$D{sr}-$E{sr})'
    )
    c_diff.font = normal_font
    c_diff.alignment = center
    c_diff.number_format = "+0.0;-0.0;0.0"
    c_diff.border = border

    c_status = ws_stat.cell(
        row=sr, column=7,
        value=f'=IF($A{sr}="","",IF(ABS($F{sr})<=1,"Rendben",IF($F{sr}>1,"Túl sok ügyelet","Túl kevés ügyelet")))'
    )
    c_status.font = normal_font
    c_status.alignment = center
    c_status.border = border

stat_last_row = last_data_row + stat_row_offset
stat_rule = FormulaRule(formula=[f'AND($A{STAT_HEADER_ROW+1}<>"",$G{STAT_HEADER_ROW+1}<>"Rendben")'], fill=ERROR_FILL)
ws_stat.conditional_formatting.add(f"A{STAT_HEADER_ROW+1}:G{stat_last_row}", stat_rule)

ws_stat.freeze_panes = f"A{STAT_HEADER_ROW+1}"

# lock the whole sheet (fully formula-driven, no manual input)
for row in ws_stat.iter_rows():
    for cell in row:
        cell.protection = openpyxl.styles.Protection(locked=True)
ws_stat.protection = SheetProtection(sheet=True, password=None, selectLockedCells=True, selectUnlockedCells=True)

# ========================================================================
# SHEET: Nyomtatási beosztás (people x days transposed view, generated by script)
# ========================================================================
ws_print = wb.create_sheet("Nyomtatási beosztás")
ws_print.sheet_view.showGridLines = False

ws_print.column_dimensions["A"].width = 26
for i in range(31):
    ws_print.column_dimensions[get_column_letter(2 + i)].width = 6
ws_print.column_dimensions["AG"].width = 16
ws_print.column_dimensions["AH"].width = 16
ws_print.column_dimensions["AI"].width = 16

ws_print["A1"] = "Nyomtatási beosztás (havi rács)"
ws_print["A1"].font = title_font
ws_print.merge_cells("A1:AI1")

ws_print["A2"] = ('Ezt a lapot a beosztás-generáló script tölti fel a Beosztás lap alapján - kód jelölés: '
                   'I=Intenzív, A=Aneszt, St=Stroke, O1/O2=osztályos, Mr=MR-altatás, el=lelépő nap, '
                   'm=rendes munkanap. A "Műtő" sor a min. 7 fős műtői létszámhoz képesti eltérést mutatja. '
                   'A Teljesített nappali óraszám a lelépős órákat is tartalmazza; havi keretesnél a teljes '
                   'ügyeleti óraszám az "Ügyeletben töltött órák" oszlopba kerül. A Túlóra a szerződéses '
                   'havi kapacitást meghaladó (vagy alatta maradó, ha negatív) órák száma - "Napi" keretnél '
                   'csak a nappali órák, folyamatos munkarendű havi keretnél az összes ledolgozott óra alapján.')
ws_print["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color="808080")
ws_print["A2"].alignment = left
ws_print.merge_cells("A2:AI2")
ws_print.row_dimensions[2].height = 40

PRINT_HEADER_ROW = 4
PRINT_START = PRINT_HEADER_ROW + 1

c_name_hdr = ws_print.cell(row=PRINT_HEADER_ROW, column=1, value="Név")
c_name_hdr.font = header_font
c_name_hdr.fill = header_fill
c_name_hdr.alignment = center
c_name_hdr.border = border
for i in range(31):
    col = 2 + i
    c_hdr = ws_print.cell(row=PRINT_HEADER_ROW, column=col, value=i + 1)
    c_hdr.font = header_font
    c_hdr.fill = header_fill
    c_hdr.alignment = center
    c_hdr.border = border

for label, col in (("Teljesített nappali óraszám", 33), ("Ügyeletben töltött órák", 34), ("Túlóra", 35)):
    c = ws_print.cell(row=PRINT_HEADER_ROW, column=col, value=label)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

for i, (name, _cat, _hrs) in enumerate(staff):
    r = PRINT_START + i
    c = ws_print.cell(row=r, column=1, value=name)
    c.font = normal_font
    c.border = border
    for col in range(2, 33):
        ws_print.cell(row=r, column=col).border = border
        ws_print.cell(row=r, column=col).alignment = center
        ws_print.cell(row=r, column=col).font = Font(name=FONT_NAME, size=9)

    c_nappali = ws_print.cell(
        row=r, column=33,
        value=f'=IF(Óraelszámolás!$B{r}="Napi",Óraelszámolás!$D{r},"")'
    )
    c_nappali.font = normal_font
    c_nappali.alignment = center
    c_nappali.number_format = '0" óra"'
    c_nappali.border = border

    c_ugyeleti = ws_print.cell(
        row=r, column=34,
        value=f'=IF(Óraelszámolás!$B{r}="Napi",Óraelszámolás!$E{r},Óraelszámolás!$F{r})'
    )
    c_ugyeleti.font = normal_font
    c_ugyeleti.alignment = center
    c_ugyeleti.number_format = '0" óra"'
    c_ugyeleti.border = border

    c_tulora = ws_print.cell(row=r, column=35)
    c_tulora.font = normal_font
    c_tulora.alignment = center
    c_tulora.number_format = '0" óra"'
    c_tulora.border = border

muto_row = PRINT_START + len(staff)
c_muto = ws_print.cell(row=muto_row, column=1, value="Műtő")
c_muto.font = bold_font
c_muto.border = border
for col in range(2, 35):
    ws_print.cell(row=muto_row, column=col).border = border
    ws_print.cell(row=muto_row, column=col).alignment = center
    ws_print.cell(row=muto_row, column=col).font = bold_font

ws_print.freeze_panes = f"B{PRINT_START}"

for row in ws_print.iter_rows():
    for cell in row:
        cell.protection = openpyxl.styles.Protection(locked=True)
ws_print.protection = SheetProtection(sheet=True, password=None, selectLockedCells=True, selectUnlockedCells=True)

# ========================================================================
# SHEET: Óraelszámolás (hour accounting per munkaidőkeret típus)
# ========================================================================
ws_hr = wb.create_sheet("Óraelszámolás")
ws_hr.sheet_view.showGridLines = False

ws_hr.column_dimensions["A"].width = 26
ws_hr.column_dimensions["B"].width = 10
ws_hr.column_dimensions["C"].width = 18
ws_hr.column_dimensions["D"].width = 20
ws_hr.column_dimensions["E"].width = 16
ws_hr.column_dimensions["F"].width = 20
ws_hr.column_dimensions["G"].width = 16
ws_hr.column_dimensions["H"].width = 14
ws_hr.column_dimensions["I"].width = 16
ws_hr.column_dimensions["J"].width = 14

ws_hr["A1"] = "Óraelszámolás (munkaidőkeret szerint)"
ws_hr["A1"].font = title_font
ws_hr.merge_cells("A1:H1")

ws_hr["A2"] = ('Napi keretes: az ügyelet napján 8 óra + (ha hétköznap - hétfő-péntek - a lelépő nap) '
               'további 8 óra kerül a kötelező munkaidőbe, a fennmaradó 16 óra az ügyeleti keretbe. '
               'Havi keretes: a teljes 24 óra egyben számít a havi keretbe, és nem lépheti túl a '
               'Dolgozók lapon megadott havi óraszámot - Havi keretes dolgozó nem kap O1/O2 beosztást.')
ws_hr["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color="808080")
ws_hr["A2"].alignment = left
ws_hr.merge_cells("A2:H2")
ws_hr.row_dimensions[2].height = 42

HR_HEADER_ROW = 4
hr_headers = ["Név", "Típus", "Ügyeletek száma (hó)", "Kötelező munkaidő (óra, Napi)",
              "Ügyeleti óra (óra, Napi)", "Havi keret teljesítés (óra, Havi)",
              "Havi óraszám kvóta", "Kvóta státusz", "Részmunkaidő kapacitás (óra)", "Kapacitás státusz"]
for j, h in enumerate(hr_headers, start=1):
    c = ws_hr.cell(row=HR_HEADER_ROW, column=j, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

hr_row_offset = HR_HEADER_ROW + 1 - row0
for r in range(row0, last_data_row + 1):
    sr = r + hr_row_offset
    c_name = ws_hr.cell(row=sr, column=1, value=f'=Dolgozók!$A{r}')
    c_name.font = normal_font
    c_name.border = border

    c_tip = ws_hr.cell(row=sr, column=2, value=f'=IF($A{sr}="","",Dolgozók!$E{r})')
    c_tip.font = normal_font
    c_tip.alignment = center
    c_tip.border = border

    c_count = ws_hr.cell(
        row=sr, column=3,
        value=f'=IF($A{sr}="","",COUNTIF(Beosztás!$C${DATA_START}:$E${DATA_END},$A{sr}))'
    )
    c_count.font = normal_font
    c_count.alignment = center
    c_count.border = border

    lelepo_parts = "+".join(
        f'SUMPRODUCT((Beosztás!${col}${DATA_START}:${col}${DATA_END}=$A{sr})*'
        f'(WEEKDAY(IF(Beosztás!$A${DATA_START}:$A${DATA_END}="",1,Beosztás!$A${DATA_START}:$A${DATA_END}+1),2)<=5))'
        for col in ("C", "D", "E")
    )
    c_kot = ws_hr.cell(
        row=sr, column=4,
        value=(f'=IF(OR($A{sr}="",$B{sr}<>"Napi"),"",'
               f'{NAPI_KOTELEZO_ORA}*$C{sr}+Dolgozók!$C{r}*({lelepo_parts}))')
    )
    c_kot.font = normal_font
    c_kot.alignment = center
    c_kot.number_format = '0" óra"'
    c_kot.border = border

    c_ugy = ws_hr.cell(
        row=sr, column=5,
        value=f'=IF(OR($A{sr}="",$B{sr}<>"Napi"),"",{NAPI_UGYELETI_ORA}*$C{sr})'
    )
    c_ugy.font = normal_font
    c_ugy.alignment = center
    c_ugy.number_format = '0" óra"'
    c_ugy.border = border

    c_havi = ws_hr.cell(
        row=sr, column=6,
        value=f'=IF(OR($A{sr}="",$B{sr}<>"Havi"),"",{HAVI_TELJES_ORA}*$C{sr})'
    )
    c_havi.font = normal_font
    c_havi.alignment = center
    c_havi.number_format = '0" óra"'
    c_havi.border = border

    c_kvota = ws_hr.cell(
        row=sr, column=7,
        value=f'=IF(OR($A{sr}="",$B{sr}<>"Havi"),"",Dolgozók!$I{r})'
    )
    c_kvota.font = normal_font
    c_kvota.alignment = center
    c_kvota.number_format = '0" óra"'
    c_kvota.border = border

    c_status = ws_hr.cell(
        row=sr, column=8,
        value=(f'=IF(OR($A{sr}="",$B{sr}<>"Havi"),"",'
               f'IF($G{sr}="","Nincs kvóta megadva!",'
               f'IF($F{sr}>$G{sr},"Túllépve! ("&($F{sr}-$G{sr})&" óra)","Rendben")))')
    )
    c_status.font = normal_font
    c_status.alignment = center
    c_status.border = border

    # Részmunkaidő - napi óraszám kapacitás: napi_munkaido × munkanapok a hónapban,
    # csak a kötelező (nem ügyeleti) órákkal összevetve.
    c_kapacitas = ws_hr.cell(
        row=sr, column=9,
        value=(f'=IF(OR($A{sr}="",$B{sr}<>"Napi",Dolgozók!$C{r}>=8),"",'
               f'Dolgozók!$C{r}*SUMPRODUCT((Beosztás!$A${DATA_START}:$A${DATA_END}<>"")*'
               f'(WEEKDAY(IF(Beosztás!$A${DATA_START}:$A${DATA_END}="",1,Beosztás!$A${DATA_START}:$A${DATA_END}),2)<=5)))')
    )
    c_kapacitas.font = normal_font
    c_kapacitas.alignment = center
    c_kapacitas.number_format = '0" óra"'
    c_kapacitas.border = border

    c_kap_status = ws_hr.cell(
        row=sr, column=10,
        value=(f'=IF($I{sr}="","",'
               f'IF($D{sr}>$I{sr}+7,"Túllépve! ("&($D{sr}-$I{sr})&" óra)",'
               f'IF($D{sr}<$I{sr},"Nem teljesíti a keretet ("&($I{sr}-$D{sr})&" óra hiányzik)",'
               f'"Rendben")))')
    )
    c_kap_status.font = normal_font
    c_kap_status.alignment = center
    c_kap_status.border = border

ws_hr.freeze_panes = f"A{HR_HEADER_ROW+1}"

hr_last_row = last_data_row + hr_row_offset
hr_rule = FormulaRule(formula=[f'OR(LEFT($H{HR_HEADER_ROW+1},4)="Túll",LEFT($J{HR_HEADER_ROW+1},4)="Túll")'], fill=ERROR_FILL)
ws_hr.conditional_formatting.add(f"A{HR_HEADER_ROW+1}:J{hr_last_row}", hr_rule)

# lock the whole sheet (fully formula-driven, no manual input)
for row in ws_hr.iter_rows():
    for cell in row:
        cell.protection = openpyxl.styles.Protection(locked=True)
ws_hr.protection = SheetProtection(sheet=True, password=None, selectLockedCells=True, selectUnlockedCells=True)

# ========================================================================
# Sheet protection (formulas locked, input cells unlocked)
# ========================================================================
# Beosztás: unlock B3 (month input) and C:E input columns
for row in ws.iter_rows():
    for cell in row:
        cell.protection = openpyxl.styles.Protection(locked=True)
ws["B3"].protection = openpyxl.styles.Protection(locked=False)
for r in range(DATA_START, DATA_END + 1):
    for col in (3, 4, 5, 6, 7):
        ws.cell(row=r, column=col).protection = openpyxl.styles.Protection(locked=False)
ws.protection = SheetProtection(sheet=True, password=None, selectLockedCells=False, selectUnlockedCells=False)
ws.protection.sheet = True
ws.protection.formatCells = False
ws.protection.formatColumns = False

# Dolgozók: unlock Név / Kategória input columns
for row in ws_staff.iter_rows():
    for cell in row:
        cell.protection = openpyxl.styles.Protection(locked=True)
for r in range(row0, last_data_row + 1):
    ws_staff.cell(row=r, column=1).protection = openpyxl.styles.Protection(locked=False)
    ws_staff.cell(row=r, column=2).protection = openpyxl.styles.Protection(locked=False)
    ws_staff.cell(row=r, column=3).protection = openpyxl.styles.Protection(locked=False)
    ws_staff.cell(row=r, column=4).protection = openpyxl.styles.Protection(locked=False)
    ws_staff.cell(row=r, column=5).protection = openpyxl.styles.Protection(locked=False)
    ws_staff.cell(row=r, column=6).protection = openpyxl.styles.Protection(locked=False)
    ws_staff.cell(row=r, column=9).protection = openpyxl.styles.Protection(locked=False)
ws_staff.protection = SheetProtection(sheet=True, password=None, selectLockedCells=False, selectUnlockedCells=False)

# Kívánságok: unlock the day-grid input cells only
for row in ws_kiv.iter_rows():
    for cell in row:
        cell.protection = openpyxl.styles.Protection(locked=True)
for r in range(KIV_START, KIV_END + 1):
    for i in range(31):
        ws_kiv.cell(row=r, column=2 + i).protection = openpyxl.styles.Protection(locked=False)
ws_kiv.protection = SheetProtection(sheet=True, password=None, selectLockedCells=False, selectUnlockedCells=False)

# ========================================================================
# Final sheet order
# ========================================================================
desired_order = ["Súgó", "Beosztás", "Nyomtatási beosztás", "Dolgozók", "Kívánságok", "Kimutatás", "Óraelszámolás", "Segéd"]
wb._sheets = [wb[name] for name in desired_order]
wb.active = 0

wb.save(KIMENET_PATH)
print("saved")
