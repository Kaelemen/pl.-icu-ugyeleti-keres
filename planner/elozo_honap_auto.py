# -*- coding: utf-8 -*-
"""Beolvassa a megosztott Google Drive mappában lévő éves beosztás-táblázatot (Google
Sheets, hónaponként külön füllel: jan, febr, márc, ápr, máj, jún, júl, aug, szept, okt,
nov, dec), és az ELŐZŐ hónap füléről megkeresi, ki volt ügyeletben annak utolsó napján
- ők lépnek le az új hónap 1-jén. Az eredményt egy JSON fájlba írja, amit a fő generáló
script beolvas és beépít a "elozo_honap_lelepok" mezőbe.

Csak OLVASÁSI jogosultságot igényel (drive.readonly) - a szolgáltatásfiók csak a vele
megosztott mappát/fájlt látja, semmi mást a Drive-on."""
import os
import io
import json
import datetime
import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
FOLDER_ID = os.environ.get("GDRIVE_FOLDER_ID", "").strip()
SA_JSON = os.environ.get("GDRIVE_SERVICE_ACCOUNT_JSON", "").strip()
# a generálandó év/hónap - ezekből számoljuk ki, melyik fület kell megnézni (előző hónap)
CEL_EV = int(os.environ.get("CEL_EV", "0") or 0)
CEL_HONAP = int(os.environ.get("CEL_HONAP", "0") or 0)
KIMENET = "elozo_honap_auto.json"

DUTY_KODOK = {"I", "A", "St"}
HONAP_FUL_NEVEK = ["", "jan", "febr", "márc", "ápr", "máj", "jún",
                   "júl", "aug", "szept", "okt", "nov", "dec"]
# Névi eltérések a Drive-táblázat és a rendszer törzsadata között - itt igazítjuk.
NEV_NORMALIZALAS = {
    "Daku Zsuzsanna": "Daku Zsuzsa",
}


def van_ugyeletkod(cellertek):
    if not cellertek:
        return False
    return any(resz in DUTY_KODOK for resz in str(cellertek).split("/"))


def talald_meg_napszam_sort(ws, max_sor=15, max_oszlop=40):
    """Megkeresi azt a sort, ahol egymást követő napszámok (1,2,3...) vannak - ez a
    fejléc-sor, amitől lefelé kezdődnek a dolgozók adatai."""
    for r in range(1, max_sor + 1):
        ertekek = [ws.cell(row=r, column=c).value for c in range(2, max_oszlop + 1)]
        szamok = [v for v in ertekek if isinstance(v, (int, float))]
        if len(szamok) >= 20 and szamok[0] == 1:
            return r
    return None


def main():
    if not SA_JSON or not CEL_HONAP:
        print("Hiányzó beállítás (kulcs/hónap) - kihagyva.")
        json.dump({"elozo_honap_lelepok": []}, open(KIMENET, "w", encoding="utf-8"))
        return

    elozo_honap = CEL_HONAP - 1 if CEL_HONAP > 1 else 12
    fulnev = HONAP_FUL_NEVEK[elozo_honap]
    print(f"Célhónap: {CEL_EV}.{CEL_HONAP:02d} - az előző hónap fülét keressük: '{fulnev}'")

    try:
        info = json.loads(SA_JSON)
        creds = service_account.Credentials.from_service_account_info(info, scopes=SCOPES)
        service = build("drive", "v3", credentials=creds)

        # A mappa-alapú keresés helyett egyszerűen az összes, a szolgáltatásfiók számára
        # látható (vele megosztott) Excel/Táblázat fájl közül a legutóbb módosítottat
        # használjuk - ez rugalmasabb, nem függ attól, pontosan hogyan lett megosztva.
        results = service.files().list(
            q="(mimeType='application/vnd.google-apps.spreadsheet' or "
              "mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet') "
              "and trashed = false",
            orderBy="modifiedTime desc",
            pageSize=10,
            fields="files(id, name, modifiedTime, mimeType)",
        ).execute()
        fajlok = results.get("files", [])
        if not fajlok:
            print("Nem található fájl a megosztott mappában - kihagyva.")
            json.dump({"elozo_honap_lelepok": []}, open(KIMENET, "w", encoding="utf-8"))
            return

        cel_fajl = fajlok[0]
        print(f"Használt fájl: {cel_fajl['name']} ({cel_fajl['mimeType']})")

        if cel_fajl["mimeType"] == "application/vnd.google-apps.spreadsheet":
            # natív Google Táblázat - exportáljuk xlsx-ként
            request = service.files().export(
                fileId=cel_fajl["id"],
                mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            # feltöltött .xlsx bináris fájl
            request = service.files().get_media(fileId=cel_fajl["id"])

        buf = io.BytesIO()
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        buf.seek(0)

        wb = openpyxl.load_workbook(buf, data_only=True)

        lap_nev = None
        for jelolt in wb.sheetnames:
            if jelolt.strip().lower() == fulnev:
                lap_nev = jelolt
                break
        if lap_nev is None:
            print(f"Nem található '{fulnev}' nevű fül a fájlban (elérhető fülek: {wb.sheetnames}) - kihagyva.")
            json.dump({"elozo_honap_lelepok": []}, open(KIMENET, "w", encoding="utf-8"))
            return

        ws = wb[lap_nev]
        print(f"Használt munkalap: {lap_nev}")

        napszam_sor = talald_meg_napszam_sort(ws)
        if napszam_sor is None:
            print("Nem található napszám-fejléc sor ezen a lapon - kihagyva.")
            json.dump({"elozo_honap_lelepok": []}, open(KIMENET, "w", encoding="utf-8"))
            return

        # az utolsó nap oszlopa: a legmagasabb egész szám a fejléc-sorban
        utolso_nap_oszlop = None
        utolso_nap_ertek = 0
        for c in range(2, 41):
            v = ws.cell(row=napszam_sor, column=c).value
            if isinstance(v, (int, float)) and v >= utolso_nap_ertek:
                utolso_nap_ertek = v
                utolso_nap_oszlop = c

        adat_kezdo_sor = napszam_sor + 2  # napszám sor + hetinap-rövidítés sor után
        lelepok = []
        for r in range(adat_kezdo_sor, adat_kezdo_sor + 30):
            nev = ws.cell(row=r, column=1).value
            if not nev or not isinstance(nev, str):
                continue
            cellertek = ws.cell(row=r, column=utolso_nap_oszlop).value
            if van_ugyeletkod(cellertek):
                lelepok.append(NEV_NORMALIZALAS.get(nev.strip(), nev.strip()))

        print(f"Az utolsó napon ({int(utolso_nap_ertek)}.) ügyeletben lévők (ők lépnek le): {lelepok}")
        json.dump({"elozo_honap_lelepok": lelepok}, open(KIMENET, "w", encoding="utf-8"), ensure_ascii=False)

    except Exception as e:
        print(f"Hiba a Drive-olvasás közben ({e}) - üres listával folytatjuk, az admin felület kézi bevitele lesz a mérvadó.")
        json.dump({"elozo_honap_lelepok": []}, open(KIMENET, "w", encoding="utf-8"))


if __name__ == "__main__":
    main()

