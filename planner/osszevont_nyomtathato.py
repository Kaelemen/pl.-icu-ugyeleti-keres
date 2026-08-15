# -*- coding: utf-8 -*-
"""Az összes legenerált beosztás-variánsból kiveszi a "Nyomtatási beosztás" lapot
(értékként, a más lapokra mutató képletek feloldva), és egyetlen közös munkafüzetbe
teszi, fülenként egy-egy változattal (Változat 1, Változat 2, ...)."""
import sys
import glob
import copy
import openpyxl

MEGTARTANDO_LAP = "Nyomtatási beosztás"

fajlok = sys.argv[2:] if len(sys.argv) > 2 else sorted(glob.glob("ICU_ugyeleti_beosztas_kimenet*_valtozat*.xlsx"))
kimenet = sys.argv[1] if len(sys.argv) > 1 else "ICU_ugyeleti_beosztas_osszes_valtozat.xlsx"

uj_wb = openpyxl.Workbook()
uj_wb.remove(uj_wb.active)

for i, fajl in enumerate(fajlok, start=1):
    # data_only=True: a más lapokra mutató képletek helyett a korábban (LibreOffice-szal)
    # újraszámolt, gyorsítótárazott ÉRTÉKEKET olvassuk ki - ezért fontos, hogy a fájl már
    # újra legyen számolva, mielőtt ez a script lefut.
    wb_ertekek = openpyxl.load_workbook(fajl, data_only=True)
    wb_stilus = openpyxl.load_workbook(fajl, data_only=False)
    if MEGTARTANDO_LAP not in wb_ertekek.sheetnames:
        print(f"Kihagyva ({fajl}): nincs '{MEGTARTANDO_LAP}' munkalap")
        continue
    forras_ertekek = wb_ertekek[MEGTARTANDO_LAP]
    forras_stilus = wb_stilus[MEGTARTANDO_LAP]

    lap_nev = MEGTARTANDO_LAP if len(fajlok) == 1 else f"Változat {i}"
    uj_lap = uj_wb.create_sheet(lap_nev)

    for row in forras_stilus.iter_rows():
        for cell in row:
            uj_cell = uj_lap.cell(row=cell.row, column=cell.column)
            ertek_cell = forras_ertekek.cell(row=cell.row, column=cell.column)
            # ha képlet volt, az újraszámolt értéket tesszük be; egyébként az eredetit
            uj_cell.value = ertek_cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else cell.value
            if cell.has_style:
                uj_cell.font = copy.copy(cell.font)
                uj_cell.fill = copy.copy(cell.fill)
                uj_cell.border = copy.copy(cell.border)
                uj_cell.alignment = copy.copy(cell.alignment)
                uj_cell.number_format = cell.number_format

    for col_letter, dim in forras_stilus.column_dimensions.items():
        uj_lap.column_dimensions[col_letter].width = dim.width
    for row_num, dim in forras_stilus.row_dimensions.items():
        uj_lap.row_dimensions[row_num].height = dim.height
    for merged in forras_stilus.merged_cells.ranges:
        uj_lap.merge_cells(str(merged))
    uj_lap.sheet_view.showGridLines = forras_stilus.sheet_view.showGridLines
    if forras_stilus.freeze_panes:
        uj_lap.freeze_panes = forras_stilus.freeze_panes

    print(f"Hozzáadva: {fajl} -> '{lap_nev}'")

uj_wb.save(kimenet)
print(f"Kész: {kimenet}")
